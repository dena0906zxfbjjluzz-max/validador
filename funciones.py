import io
import os
import re
import hashlib
import datetime
import sqlite3
import json
import urllib.error
import urllib.parse
import urllib.request
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# Base SQLite local genérica (no atada a una planta específica).
# Si el archivo no existe, se crea al conectar e `inicializar_base_datos()` genera
# todas las tablas de auditoría con la estructura limpia actual.
_DB_DIR = os.path.dirname(os.path.abspath(__file__))
DB_NAME = "planta_calidad_prod.db"
DB_PATH = os.path.join(_DB_DIR, DB_NAME)


def _conectar_db():
    """Abre (o crea) la base genérica planta_calidad_prod.db."""
    os.makedirs(_DB_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def inicializar_base_datos():
    """
    Crea la BD y tablas de auditoría si no existen.
    Al usar una base nueva (p. ej. tras el rename genérico), el esquema se
    construye desde cero con la estructura limpia actual.
    """
    es_nueva = not os.path.exists(DB_PATH)
    conn = _conectar_db()
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS historial_sesion (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hora TEXT,
            archivo TEXT,
            registros INTEGER,
            confiabilidad TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bitacora_cambios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha_hora TEXT,
            fila_indice INTEGER,
            columna TEXT,
            valor_anterior TEXT,
            nuevo_valor TEXT,
            inspector TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS control_frio (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            camara TEXT,
            temperatura REAL,
            hora_registro TEXT,
            estado TEXT,
            inspector TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS contenedores_despacho (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            booking TEXT,
            contenedor TEXT,
            precinto_linea TEXT,
            precinto_senasa TEXT,
            destino TEXT,
            estado TEXT
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS historial_reportes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha_hora TEXT NOT NULL,
            lote TEXT NOT NULL,
            hash_sha256 TEXT NOT NULL UNIQUE,
            responsable TEXT NOT NULL,
            archivo TEXT,
            producto TEXT,
            registros INTEGER,
            firma_ecc TEXT,
            llave_publica TEXT,
            mensaje TEXT,
            modo_firma TEXT,
            backend TEXT
        )
    """)
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_historial_reportes_fecha ON historial_reportes(fecha_hora DESC)"
    )
    cursor.execute(
        "CREATE INDEX IF NOT EXISTS idx_bitacora_fecha ON bitacora_cambios(fecha_hora)"
    )

    # Solo en bases antiguas con esquema incompleto (no aplica a archivos nuevos)
    if not es_nueva:
        try:
            cursor.execute("ALTER TABLE control_frio ADD COLUMN inspector TEXT")
        except sqlite3.OperationalError:
            pass
        try:
            cursor.execute(
                "DELETE FROM bitacora_cambios WHERE datetime(fecha_hora) < datetime('now', '-7 days')"
            )
        except sqlite3.OperationalError:
            pass

    conn.commit()
    conn.close()
    return DB_PATH


def calcular_hash_reporte(mensaje: str, firma: str, llave_publica: str = "") -> str:
    """Huella SHA-256 del sello ECC (mensaje + firma + llave pública)."""
    payload = f"{mensaje}|{firma}|{llave_publica}".encode("utf-8")
    return hashlib.sha256(payload).hexdigest()

# Alias de columnas frecuentes en packing lists agroindustriales
ALIAS_COLUMNAS = {
    "id_unidad": ["CAJA", "SSCC", "PALLET", "CODIGO", "BARCODE", "EAN", "GTIN", "ID"],
    "lote": ["LOTE"],
    "fundo": ["FUNDO", "PARCELA", "CAMPO", "ORIGEN", "SECTOR"],
    "productor": ["PRODUCTOR", "PROVEEDOR", "AGRICULTOR"],
    "turno": ["TURNO", "LINEA", "PROCESO", "PACKING"],
    "peso": ["PESO"],
    "cosecha": ["COSECHA", "HARVEST", "FECHA_COSECHA", "F_COSECHA"],
    "lmr": ["LMR", "RESIDUO", "ANALISIS", "FITOSANITARIO"],
    "calibre": ["CALIBRE"],
    "categoria": ["CATEGORIA", "CAT"],
}


def encontrar_columna(df, keywords):
    """Devuelve el nombre real de la primera columna cuyo nombre contenga alguno de los keywords."""
    for col in df.columns:
        col_up = str(col).upper()
        for kw in keywords:
            if kw.upper() in col_up:
                return col
    return None


def mapear_columnas_trazabilidad(df):
    return {clave: encontrar_columna(df, aliases) for clave, aliases in ALIAS_COLUMNAS.items()}


CAMPOS_MAPEO_UI = (
    "id_unidad",
    "lote",
    "peso",
    "calibre",
    "fundo",
    "productor",
    "turno",
    "cosecha",
    "lmr",
    "categoria",
)
CAMPOS_CRITICOS_PACKING = ("lote", "peso", "calibre")


def resolver_mapa_columnas(df, mapeo_manual: dict | None = None) -> dict:
    """
    Combina detección automática por alias con un mapeo manual del usuario.
    En mapeo_manual: valor None o '' deja el auto; '(ninguna)' fuerza sin columna.
    """
    auto = mapear_columnas_trazabilidad(df)
    if not mapeo_manual:
        return auto
    out = dict(auto)
    cols = set(str(c) for c in df.columns)
    for clave, elegido in mapeo_manual.items():
        if clave not in out:
            continue
        if elegido is None or elegido == "" or elegido == "(auto)":
            continue
        if elegido == "(ninguna)":
            out[clave] = None
            continue
        if str(elegido) in cols:
            out[clave] = str(elegido)
    return out


def campos_criticos_sin_mapear(mapa: dict | None) -> list[str]:
    """Nombres legibles de LOTE/PESO/CALIBRE aún sin columna asignada."""
    mapa = mapa or {}
    return [k.upper() for k in CAMPOS_CRITICOS_PACKING if not mapa.get(k)]


def _valor_celda(fila, col, default="N/D"):
    if col is None or col not in fila.index:
        return default
    val = fila[col]
    if pd.isna(val) or str(val).strip() in ("", "-"):
        return default
    return str(val).strip()


def buscar_registros_por_codigo(df, codigo):
    """Busca filas cuyo ID (caja/pallet/sscc/código) o cualquier celda coincida con el código."""
    codigo = str(codigo).strip()
    if not codigo or df.empty:
        return df.iloc[0:0].copy()

    cols_mapa = mapear_columnas_trazabilidad(df)
    cols_prioridad = [
        c for c in [
            cols_mapa["id_unidad"],
            cols_mapa["lote"],
            encontrar_columna(df, ["CONTENEDOR", "BOOKING"]),
        ] if c is not None
    ]
    # Evitar duplicados manteniendo orden
    cols_busqueda = list(dict.fromkeys(cols_prioridad + list(df.columns)))

    mask = pd.Series(False, index=df.index)
    for col in cols_busqueda:
        serie = df[col].astype(str).str.strip()
        mask = mask | serie.str.fullmatch(codigo, case=False, na=False)
        if mask.any():
            break

    if not mask.any():
        # Búsqueda parcial solo en columnas de identificación
        for col in cols_prioridad:
            serie = df[col].astype(str).str.strip()
            mask = mask | serie.str.contains(codigo, case=False, na=False, regex=False)

    return df.loc[mask].copy()


def armar_arbol_trazabilidad(fila, cols_mapa, inspector, codigo_buscado):
    """Construye un diccionario legible del árbol genealógico a partir de una fila real."""
    return {
        "codigo": codigo_buscado,
        "fundo": _valor_celda(fila, cols_mapa["fundo"], "No informado en archivo"),
        "productor": _valor_celda(fila, cols_mapa["productor"], "No informado en archivo"),
        "lote": _valor_celda(fila, cols_mapa["lote"]),
        "turno": _valor_celda(fila, cols_mapa["turno"], "No informado en archivo"),
        "cosecha": _valor_celda(fila, cols_mapa["cosecha"], "No informado en archivo"),
        "peso": _valor_celda(fila, cols_mapa["peso"]),
        "calibre": _valor_celda(fila, cols_mapa["calibre"]),
        "lmr": _valor_celda(fila, cols_mapa["lmr"], "Sin dato LMR en archivo"),
        "inspector": inspector,
        "fila_indice": int(fila.name) if fila.name is not None else None,
    }


def interpretar_estado_lmr(texto):
    """Clasifica un resultado LMR textual en conforme / alerta / rechazado."""
    t = str(texto).strip().upper()
    if not t or t in ("N/D", "-", "SIN DATO LMR EN ARCHIVO", "NAN"):
        return "sin_dato"
    if any(k in t for k in ["RECHAZ", "SUPERA", "FAIL", "NO CONFORME", "BLOQUE"]):
        return "rechazado"
    if any(k in t for k in ["ALERTA", "CERCANO", "CUARENTENA", "PENDIENTE", "WARNING"]):
        return "alerta"
    if any(k in t for k in ["CONFORME", "APROB", "OK", "BAJO", "PASS", "CUMPLE"]):
        return "conforme"
    return "sin_dato"


def registrar_peso_ultima_fila(df, peso):
    """Inyecta el peso capturado en la última fila de la columna PESO (si existe)."""
    df_out = df.copy()
    col_peso = encontrar_columna(df_out, ALIAS_COLUMNAS["peso"])
    if col_peso is None or df_out.empty:
        return df_out, False, col_peso
    df_out.iloc[-1, df_out.columns.get_loc(col_peso)] = str(peso)
    return df_out, True, col_peso


# Tabla destino en Supabase: public.historial_reportes
SUPABASE_TABLA_SELLOS = "historial_reportes"
# Solo estas columnas (minúsculas exactas, sin extras)
SUPABASE_CAMPOS_SELLO = ("fecha", "lote", "hash_sha256", "inspector")


def _supabase_config():
    """
    Lee SUPABASE_URL y SUPABASE_KEY desde st.secrets.

    Orden de búsqueda:
      1) st.secrets["credenciales"]["SUPABASE_URL"] / ["SUPABASE_KEY"]  (Cloud actual)
      2) st.secrets["SUPABASE_URL"] / ["SUPABASE_KEY"]  (raíz)
      3) st.secrets["supabase"]["url"] / ["key"]

    Returns: (url, key, error_si_falla)
    """
    try:
        import streamlit as st
    except Exception as e:
        return None, None, f"Streamlit no disponible para secrets: {e}"

    try:
        secrets = st.secrets
    except Exception as e:
        return None, None, f"No se pudieron leer secrets: {e}"

    def _limpiar(val):
        if val is None:
            return None
        s = str(val).strip().strip('"').strip("'")
        return s or None

    def _desde_bloque(bloque, *nombres):
        if bloque is None:
            return None
        for nombre in nombres:
            try:
                v = bloque[nombre]
            except Exception:
                try:
                    v = bloque.get(nombre)  # type: ignore[attr-defined]
                except Exception:
                    continue
            limpio = _limpiar(v)
            if limpio:
                return limpio
        return None

    url = None
    key = None
    origen = None

    # 1) Preferido: dentro de [credenciales]  (diagnóstico: solo existe esta clave raíz)
    try:
        creds = secrets["credenciales"]
        url = _desde_bloque(
            creds,
            "SUPABASE_URL",
            "supabase_url",
            "URL",
            "url",
        )
        key = _desde_bloque(
            creds,
            "SUPABASE_KEY",
            "SUPABASE_SERVICE_KEY",
            "supabase_key",
            "service_key",
            "KEY",
            "key",
            "anon_key",
        )
        if url and key:
            origen = "st.secrets['credenciales']"
    except Exception:
        creds = None

    # 2) Raíz
    if not url or not key:
        try:
            url = url or _limpiar(secrets["SUPABASE_URL"])
        except Exception:
            pass
        try:
            url = url or _limpiar(secrets["supabase_url"])
        except Exception:
            pass
        try:
            key = key or _limpiar(secrets["SUPABASE_KEY"])
        except Exception:
            pass
        try:
            key = key or _limpiar(secrets["SUPABASE_SERVICE_KEY"])
        except Exception:
            pass
        try:
            key = key or _limpiar(secrets["supabase_key"])
        except Exception:
            pass
        if url and key and origen is None:
            origen = "st.secrets (raíz)"

    # 3) Sección [supabase]
    if not url or not key:
        try:
            bloque = secrets["supabase"]
            url = url or _desde_bloque(bloque, "url", "URL", "SUPABASE_URL")
            key = key or _desde_bloque(
                bloque, "key", "KEY", "SUPABASE_KEY", "service_key", "anon_key"
            )
            if url and key and origen is None:
                origen = "st.secrets['supabase']"
        except Exception:
            pass

    if not url or not key:
        extras = ""
        try:
            claves = sorted(str(k) for k in secrets.keys())
            extras = f" Claves en secrets: {claves}."
        except Exception:
            pass
        detalle_creds = ""
        try:
            creds = secrets["credenciales"]
            sub = sorted(str(k) for k in creds.keys())
            detalle_creds = f" Claves dentro de credenciales: {sub}."
        except Exception:
            detalle_creds = (
                " No se encontraron subclaves en [credenciales]. "
                "Defina SUPABASE_URL y SUPABASE_KEY dentro de ese bloque."
            )
        return (
            None,
            None,
            "Faltan SUPABASE_URL o SUPABASE_KEY. "
            "Úselas como st.secrets['credenciales']['SUPABASE_URL'] y "
            "st.secrets['credenciales']['SUPABASE_KEY']."
            + extras
            + detalle_creds,
        )

    if not url.startswith("http"):
        return None, None, f"SUPABASE_URL inválida (debe empezar con https://): {url[:40]}..."

    return url.rstrip("/"), key, None


def enviar_sello_a_supabase(
    fecha,
    lote,
    hash_sha256,
    inspector,
    tabla=None,
    timeout=15.0,
):
    """
    POST REST a public.historial_reportes con EXACTAMENTE:
      { "fecha", "lote", "hash_sha256", "inspector" }
    en minúsculas. Sin campos extra (fecha_hora/responsable ya no se envían).

    Returns dict: ok, configurado, status, mensaje, payload, endpoint
    """
    url, key, err_cfg = _supabase_config()
    if err_cfg or not url or not key:
        return {
            "ok": False,
            "configurado": False,
            "status": None,
            "mensaje": err_cfg or "Supabase no configurado",
            "payload": {},
            "endpoint": None,
        }

    tabla_destino = (tabla or SUPABASE_TABLA_SELLOS).strip() or SUPABASE_TABLA_SELLOS
    # PostgREST: /rest/v1/historial_reportes → public.historial_reportes
    endpoint = f"{url}/rest/v1/{tabla_destino}"

    # Payload estricto — solo las 4 columnas de la tabla
    payload = {
        "fecha": str(fecha or "").strip(),
        "lote": str(lote or "N/D").strip(),
        "hash_sha256": str(hash_sha256 or "").strip(),
        "inspector": str(inspector or "N/D").strip(),
    }

    # Validación local de campos vacíos
    vacios = [c for c in SUPABASE_CAMPOS_SELLO if not payload.get(c)]
    if vacios:
        return {
            "ok": False,
            "configurado": True,
            "status": None,
            "mensaje": f"Campos vacíos, no se envió a Supabase: {vacios}",
            "payload": payload,
            "endpoint": endpoint,
        }

    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Accept": "application/json",
        # Devuelve la fila insertada para confirmar que no falló en silencio
        "Prefer": "return=representation",
    }

    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(endpoint, data=body, method="POST", headers=headers)

    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            status = getattr(resp, "status", 200) or 200
            raw = resp.read().decode("utf-8", errors="replace")
            if status in (200, 201, 204):
                # return=representation suele devolver un array JSON
                return {
                    "ok": True,
                    "configurado": True,
                    "status": status,
                    "mensaje": (
                        f"INSERT OK en public.{tabla_destino} "
                        f"(HTTP {status}). Respuesta: {(raw or '[]')[:200]}"
                    ),
                    "payload": payload,
                    "endpoint": endpoint,
                    "respuesta": (raw or "")[:500],
                }
            return {
                "ok": False,
                "configurado": True,
                "status": status,
                "mensaje": f"Supabase HTTP {status}: {(raw or 'sin cuerpo')[:400]}",
                "payload": payload,
                "endpoint": endpoint,
            }
    except urllib.error.HTTPError as e:
        try:
            detalle = e.read().decode("utf-8", errors="replace")[:600]
        except Exception:
            detalle = str(e)
        hint = ""
        if e.code in (401, 403):
            hint = (
                " | Pista: use la service_role key o cree una policy RLS de INSERT "
                "en public.historial_reportes."
            )
        elif e.code == 404:
            hint = " | Pista: la tabla public.historial_reportes no existe o el schema no es public."
        elif e.code == 409:
            # Hash único ya en la nube: no es un fallo operativo
            return {
                "ok": True,
                "ya_existia": True,
                "configurado": True,
                "status": 409,
                "mensaje": (
                    "Sello ya registrado en Supabase (hash_sha256 único). "
                    "No se duplicó el registro."
                ),
                "payload": payload,
                "endpoint": endpoint,
            }
        return {
            "ok": False,
            "configurado": True,
            "status": e.code,
            "mensaje": f"Error HTTP {e.code} POST {endpoint}: {detalle}{hint}",
            "payload": payload,
            "endpoint": endpoint,
        }
    except urllib.error.URLError as e:
        return {
            "ok": False,
            "configurado": True,
            "status": None,
            "mensaje": f"Error de red al conectar con Supabase ({endpoint}): {e.reason}",
            "payload": payload,
            "endpoint": endpoint,
        }
    except Exception as e:
        return {
            "ok": False,
            "configurado": True,
            "status": None,
            "mensaje": f"Excepción al enviar a Supabase: {type(e).__name__}: {e}",
            "payload": payload,
            "endpoint": endpoint,
        }


def _replicar_reporte_supabase(registro: dict):
    """Wrapper legacy. Preferir enviar_sello_a_supabase."""
    resultado = enviar_sello_a_supabase(
        fecha=str(registro.get("fecha") or registro.get("fecha_hora") or ""),
        lote=str(registro.get("lote") or "N/D"),
        hash_sha256=str(registro.get("hash_sha256") or ""),
        inspector=str(registro.get("inspector") or registro.get("responsable") or "N/D"),
    )
    if not resultado.get("configurado"):
        return None
    if resultado.get("ok"):
        return "ok"
    return resultado.get("mensaje") or "error Supabase"


def guardar_reporte_historico(
    fecha_hora: str,
    lote: str,
    hash_sha256: str,
    responsable: str,
    archivo: str = "",
    producto: str = "",
    registros: int = 0,
    firma_ecc: str = "",
    llave_publica: str = "",
    mensaje: str = "",
    modo_firma: str = "",
    backend: str = "",
) -> dict:
    """
    SQLite local + copia REST a Supabase (fecha, lote, hash_sha256, inspector).
    """
    inicializar_base_datos()
    conn = _conectar_db()
    cursor = conn.cursor()
    cursor.execute(
        "SELECT id FROM historial_reportes WHERE hash_sha256 = ?", (hash_sha256,)
    )
    existente = cursor.fetchone()

    ya_existia = bool(existente)
    new_id = existente[0] if existente else None

    if not ya_existia:
        cursor.execute(
            """
            INSERT INTO historial_reportes (
                fecha_hora, lote, hash_sha256, responsable, archivo, producto,
                registros, firma_ecc, llave_publica, mensaje, modo_firma, backend
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                fecha_hora,
                lote or "N/D",
                hash_sha256,
                responsable or "N/D",
                archivo,
                producto,
                int(registros or 0),
                firma_ecc,
                llave_publica,
                mensaje,
                modo_firma,
                backend,
            ),
        )
        new_id = cursor.lastrowid
        conn.commit()
    conn.close()

    try:
        supabase_result = enviar_sello_a_supabase(
            fecha=fecha_hora,
            lote=lote or "N/D",
            hash_sha256=hash_sha256,
            inspector=responsable or "N/D",
        )
    except Exception as e:
        supabase_result = {
            "ok": False,
            "configurado": True,
            "status": None,
            "mensaje": f"{type(e).__name__}: {e}",
            "payload": {
                "fecha": fecha_hora,
                "lote": lote or "N/D",
                "hash_sha256": hash_sha256,
                "inspector": responsable or "N/D",
            },
            "endpoint": None,
        }

    return {
        "guardado": not ya_existia,
        "ya_existia": ya_existia,
        "id": new_id,
        "hash": hash_sha256,
        "supabase": "ok" if supabase_result.get("ok") else supabase_result.get("mensaje"),
        "supabase_detalle": supabase_result,
    }


def cargar_historial_reportes_db(limite: int = 200) -> list:
    inicializar_base_datos()
    conn = _conectar_db()
    df = pd.read_sql_query(
        """
        SELECT
            fecha_hora AS 'Fecha',
            lote AS 'Lote',
            hash_sha256 AS 'Hash',
            responsable AS 'Responsable',
            archivo AS 'Archivo',
            producto AS 'Producto',
            registros AS 'Registros',
            modo_firma AS 'Modo',
            backend AS 'Backend'
        FROM historial_reportes
        ORDER BY id DESC
        LIMIT ?
        """,
        conn,
        params=(int(limite),),
    )
    conn.close()
    return df.to_dict("records")


def buscar_reporte_por_hash(hash_sha256: str) -> dict | None:
    inicializar_base_datos()
    conn = _conectar_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT fecha_hora, lote, hash_sha256, responsable, archivo, producto,
               registros, firma_ecc, llave_publica, mensaje, modo_firma, backend
        FROM historial_reportes
        WHERE hash_sha256 = ?
        LIMIT 1
        """,
        (hash_sha256.strip().lower(),),
    )
    row = cursor.fetchone()
    conn.close()
    if not row:
        # también buscar sin forzar lower en caso de mayúsculas mixtas
        conn = _conectar_db()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT fecha_hora, lote, hash_sha256, responsable, archivo, producto,
                   registros, firma_ecc, llave_publica, mensaje, modo_firma, backend
            FROM historial_reportes
            WHERE lower(hash_sha256) = lower(?)
            LIMIT 1
            """,
            (hash_sha256.strip(),),
        )
        row = cursor.fetchone()
        conn.close()
    if not row:
        return None
    keys = [
        "fecha_hora", "lote", "hash_sha256", "responsable", "archivo", "producto",
        "registros", "firma_ecc", "llave_publica", "mensaje", "modo_firma", "backend",
    ]
    return dict(zip(keys, row))


# ─── Módulo Escaneo QR → historial_reportes (Supabase) ───────────────────────

_RE_HASH_SHA256 = re.compile(r"\b([a-fA-F0-9]{64})\b")
_RE_LOTE_KV = re.compile(
    r"(?:lote|lot|id_lote|lote_id)\s*[:=]\s*([^\s,;|]+)",
    re.IGNORECASE,
)
_RE_HASH_KV = re.compile(
    r"(?:hash_sha256|hash|sha256|sello)\s*[:=]\s*([a-fA-F0-9]{64})",
    re.IGNORECASE,
)


def construir_payload_qr(lote: str, hash_sha256: str, inspector: str = "") -> str:
    """Texto canónico del QR del pallet (JSON legible para pyzbar)."""
    payload = {
        "lote": str(lote or "").strip() or "N/D",
        "hash_sha256": str(hash_sha256 or "").strip().lower(),
    }
    if inspector:
        payload["inspector"] = str(inspector).strip()
    return json.dumps(payload, ensure_ascii=False)


def parsear_payload_qr(texto_qr: str) -> dict:
    """
    Extrae lote y hash_sha256 del QR.
    Formatos aceptados: JSON, clave=valor, pipe, hash hex puro, URL con query.
    """
    crudo = (texto_qr or "").strip()
    result = {
        "texto": crudo,
        "lote": None,
        "hash_sha256": None,
    }
    if not crudo:
        return result

    # 1) JSON
    if crudo.startswith("{") and crudo.endswith("}"):
        try:
            obj = json.loads(crudo)
            if isinstance(obj, dict):
                for k_hash in ("hash_sha256", "hash", "sha256", "sello"):
                    if obj.get(k_hash):
                        result["hash_sha256"] = str(obj[k_hash]).strip().lower()
                        break
                for k_lote in ("lote", "lot", "id_lote", "lote_id", "id"):
                    if obj.get(k_lote):
                        result["lote"] = str(obj[k_lote]).strip()
                        break
                if result["hash_sha256"] or result["lote"]:
                    return result
        except json.JSONDecodeError:
            pass

    # 2) Query string / URL
    try:
        parsed = urllib.parse.urlparse(crudo)
        qs = urllib.parse.parse_qs(parsed.query)
        for k, vals in qs.items():
            if not vals:
                continue
            kl = k.lower()
            if kl in ("hash_sha256", "hash", "sha256") and not result["hash_sha256"]:
                result["hash_sha256"] = str(vals[0]).strip().lower()
            if kl in ("lote", "lot", "id_lote") and not result["lote"]:
                result["lote"] = str(vals[0]).strip()
        if result["hash_sha256"] or result["lote"]:
            return result
    except Exception:
        pass

    # 3) Hash hex puro
    if re.fullmatch(r"[a-fA-F0-9]{64}", crudo):
        result["hash_sha256"] = crudo.lower()
        return result

    # 4) lote|hash o hash|lote
    if "|" in crudo:
        partes = [p.strip() for p in crudo.split("|") if p.strip()]
        for p in partes:
            if re.fullmatch(r"[a-fA-F0-9]{64}", p):
                result["hash_sha256"] = p.lower()
            elif not result["lote"] and p.lower() not in ("n/d", "nan"):
                result["lote"] = p
        if result["hash_sha256"] or result["lote"]:
            return result

    # 5) Clave: valor en el texto
    m_h = _RE_HASH_KV.search(crudo)
    if m_h:
        result["hash_sha256"] = m_h.group(1).lower()
    m_l = _RE_LOTE_KV.search(crudo)
    if m_l:
        result["lote"] = m_l.group(1).strip()
    if not result["hash_sha256"]:
        m_any = _RE_HASH_SHA256.search(crudo)
        if m_any:
            result["hash_sha256"] = m_any.group(1).lower()

    # 6) Si no hay hash, el texto completo puede ser el ID de lote
    if not result["hash_sha256"] and not result["lote"] and crudo and len(crudo) < 120:
        result["lote"] = crudo

    return result


def decodificar_qr_desde_imagen(imagen_bytes: bytes) -> dict:
    """
    Decodifica QR/códigos desde bytes de imagen (st.camera_input / upload).
    Usa Pillow + pyzbar.
    """
    if not imagen_bytes:
        return {
            "ok": False,
            "textos": [],
            "mensaje": "No hay imagen para decodificar.",
        }

    try:
        from PIL import Image
    except ImportError:
        return {
            "ok": False,
            "textos": [],
            "mensaje": "Falta Pillow. Añada `Pillow` a requirements.txt.",
        }

    try:
        from pyzbar.pyzbar import decode as pyzbar_decode
    except ImportError:
        return {
            "ok": False,
            "textos": [],
            "mensaje": (
                "Falta pyzbar (o libzbar del sistema). "
                "Instale `pyzbar` y el paquete del SO `libzbar0`."
            ),
        }

    try:
        img = Image.open(io.BytesIO(imagen_bytes))
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        codigos = pyzbar_decode(img)
    except Exception as e:
        return {
            "ok": False,
            "textos": [],
            "mensaje": f"Error al decodificar imagen: {type(e).__name__}: {e}",
        }

    textos = []
    detalles = []
    for c in codigos or []:
        try:
            t = c.data.decode("utf-8", errors="replace").strip()
        except Exception:
            t = str(getattr(c, "data", "")).strip()
        if t:
            textos.append(t)
            detalles.append(
                {
                    "tipo": str(getattr(c, "type", "") or "QR"),
                    "texto": t,
                }
            )

    if not textos:
        return {
            "ok": False,
            "textos": [],
            "detalles": [],
            "mensaje": "No se detectó ningún código QR en la imagen. Encuadre el código y capture de nuevo.",
        }

    return {
        "ok": True,
        "textos": textos,
        "detalles": detalles,
        "texto": textos[0],
        "mensaje": f"QR leído ({len(textos)} código(s)).",
    }


def consultar_historial_reportes_supabase(
    hash_sha256: str | None = None,
    lote: str | None = None,
    timeout: float = 15.0,
    limite: int = 20,
) -> dict:
    """
    GET REST equivalente a:
      supabase.table('historial_reportes').select('*').eq('hash_sha256', ...).eq('lote', ...)

    Filtra por hash y/o lote en public.historial_reportes.
    """
    url, key, err_cfg = _supabase_config()
    if err_cfg or not url or not key:
        return {
            "ok": False,
            "configurado": False,
            "encontrado": False,
            "filas": [],
            "mensaje": err_cfg or "Supabase no configurado",
            "endpoint": None,
        }

    params = {
        "select": "fecha,lote,hash_sha256,inspector",
        "limit": str(int(limite)),
        "order": "fecha.desc",
    }
    h = (hash_sha256 or "").strip().lower()
    l = (lote or "").strip()
    if h:
        params["hash_sha256"] = f"eq.{h}"
    if l:
        params["lote"] = f"eq.{l}"

    if "hash_sha256" not in params and "lote" not in params:
        return {
            "ok": False,
            "configurado": True,
            "encontrado": False,
            "filas": [],
            "mensaje": "Se requiere hash_sha256 o lote para consultar historial_reportes.",
            "endpoint": None,
        }

    endpoint = f"{url}/rest/v1/{SUPABASE_TABLA_SELLOS}?{urllib.parse.urlencode(params)}"
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Accept": "application/json",
    }

    try:
        req = urllib.request.Request(endpoint, method="GET", headers=headers)
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            status = getattr(resp, "status", 200) or 200
            raw = resp.read().decode("utf-8", errors="replace")
            filas = []
            if raw.strip():
                try:
                    data = json.loads(raw)
                    if isinstance(data, list):
                        filas = data
                    elif isinstance(data, dict):
                        filas = [data]
                except json.JSONDecodeError:
                    return {
                        "ok": False,
                        "configurado": True,
                        "encontrado": False,
                        "filas": [],
                        "mensaje": f"Respuesta no JSON de Supabase (HTTP {status}): {raw[:200]}",
                        "endpoint": endpoint,
                    }
            return {
                "ok": True,
                "configurado": True,
                "encontrado": len(filas) > 0,
                "filas": filas,
                "status": status,
                "mensaje": (
                    f"Consulta OK public.{SUPABASE_TABLA_SELLOS}: {len(filas)} fila(s)."
                ),
                "endpoint": endpoint,
            }
    except urllib.error.HTTPError as e:
        try:
            detalle = e.read().decode("utf-8", errors="replace")[:400]
        except Exception:
            detalle = str(e)
        return {
            "ok": False,
            "configurado": True,
            "encontrado": False,
            "filas": [],
            "status": e.code,
            "mensaje": f"HTTP {e.code} GET historial_reportes: {detalle}",
            "endpoint": endpoint,
        }
    except Exception as e:
        return {
            "ok": False,
            "configurado": True,
            "encontrado": False,
            "filas": [],
            "mensaje": f"{type(e).__name__}: {e}",
            "endpoint": endpoint,
        }


def validar_pallet_por_qr(imagen_bytes: bytes = None, texto_qr: str = None) -> dict:
    """
    Pipeline ERP: imagen/cámara → pyzbar → lote/hash → Supabase historial_reportes.

    Si el QR contiene un hash_sha256 registrado en la nube → verificado.
    Si solo hay lote, valida existencia del lote en la tabla remota.
    """
    textos_qr = []
    decode_info = None

    if imagen_bytes:
        decode_info = decodificar_qr_desde_imagen(imagen_bytes)
        if not decode_info.get("ok"):
            return {
                "verificado": False,
                "tipo_ui": "error",
                "mensaje_ui": decode_info.get("mensaje") or "No se pudo leer el QR.",
                "payload": None,
                "supabase": None,
                "decode": decode_info,
            }
        textos_qr = list(decode_info.get("textos") or [])
    elif texto_qr and str(texto_qr).strip():
        textos_qr = [str(texto_qr).strip()]
    else:
        return {
            "verificado": False,
            "tipo_ui": "error",
            "mensaje_ui": "Capture una foto del QR o ingrese el texto decodificado.",
            "payload": None,
            "supabase": None,
            "decode": None,
        }

    # Priorizar payload con hash; si hay varios QR, intentar todos
    payloads = [parsear_payload_qr(t) for t in textos_qr]
    payload = None
    for p in payloads:
        if p.get("hash_sha256"):
            payload = p
            break
    if payload is None:
        payload = payloads[0] if payloads else parsear_payload_qr(textos_qr[0])

    hash_q = payload.get("hash_sha256")
    lote_q = payload.get("lote")

    # Consulta automática a public.historial_reportes
    if hash_q:
        sb = consultar_historial_reportes_supabase(hash_sha256=hash_q, lote=None)
        # Si no hay fila por hash y hay lote, revalidar enlace hash+lote
        if not sb.get("encontrado") and lote_q:
            sb_lote = consultar_historial_reportes_supabase(hash_sha256=hash_q, lote=lote_q)
            if sb_lote.get("encontrado"):
                sb = sb_lote
    elif lote_q:
        sb = consultar_historial_reportes_supabase(hash_sha256=None, lote=lote_q)
    else:
        return {
            "verificado": False,
            "tipo_ui": "error",
            "mensaje_ui": (
                "🚨 El QR no contiene un ID de lote ni un hash_sha256 reconocible."
            ),
            "payload": payload,
            "supabase": None,
            "decode": decode_info,
        }

    if not sb.get("configurado"):
        # Fallback local SQLite si Supabase no está configurado
        local = None
        if hash_q:
            local = buscar_reporte_por_hash(hash_q)
        if local:
            return {
                "verificado": True,
                "tipo_ui": "success",
                "mensaje_ui": (
                    "✅ Pallet verificado en historial local (SQLite). "
                    "Configure Supabase para validación en la nube."
                ),
                "payload": payload,
                "supabase": sb,
                "registro": {
                    "fecha": local.get("fecha_hora"),
                    "lote": local.get("lote"),
                    "hash_sha256": local.get("hash_sha256"),
                    "inspector": local.get("responsable"),
                },
                "decode": decode_info,
            }
        return {
            "verificado": False,
            "tipo_ui": "error",
            "mensaje_ui": sb.get("mensaje") or "Supabase no configurado.",
            "payload": payload,
            "supabase": sb,
            "decode": decode_info,
        }

    if not sb.get("ok"):
        return {
            "verificado": False,
            "tipo_ui": "error",
            "mensaje_ui": f"🚨 Error al consultar Supabase: {sb.get('mensaje')}",
            "payload": payload,
            "supabase": sb,
            "decode": decode_info,
        }

    if sb.get("encontrado"):
        filas = sb.get("filas") or []
        # Si el QR traía hash, debe coincidir exactamente con alguna fila
        if hash_q:
            match = any(
                str(f.get("hash_sha256") or "").strip().lower() == hash_q for f in filas
            )
            if not match:
                return {
                    "verificado": False,
                    "tipo_ui": "error",
                    "mensaje_ui": (
                        "🚨 ALERTA: el hash del QR no coincide con ningún sello "
                        "registrado en Supabase (historial_reportes)."
                    ),
                    "payload": payload,
                    "supabase": sb,
                    "decode": decode_info,
                }
        registro = filas[0] if filas else {}
        return {
            "verificado": True,
            "tipo_ui": "success",
            "mensaje_ui": (
                "✅ Pallet verificado de forma segura mediante QR en Supabase."
            ),
            "payload": payload,
            "supabase": sb,
            "registro": registro,
            "decode": decode_info,
        }

    # No encontrado en la nube
    if hash_q:
        msg = (
            "🚨 ALERTA: el hash del QR no está registrado en Supabase "
            f"(historial_reportes). Lote leído: {lote_q or 'N/D'}."
        )
    else:
        msg = (
            f"🚨 ALERTA: el lote `{lote_q}` no tiene reportes sellados "
            "en Supabase (historial_reportes)."
        )
    return {
        "verificado": False,
        "tipo_ui": "error",
        "mensaje_ui": msg,
        "payload": payload,
        "supabase": sb,
        "decode": decode_info,
    }


def procesar_escaneo_qr_camara(imagen_bytes: bytes) -> dict:
    """
    Entrada desde st.camera_input: decodifica el QR y valida en Supabase.
    Alias funcional del pipeline ERP de pallet.
    """
    return validar_pallet_por_qr(imagen_bytes=imagen_bytes)


def guardar_historial_db(hora, archivo, registros, confiabilidad):
    conn = _conectar_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM historial_sesion WHERE archivo = ?", (archivo,))
    if not cursor.fetchone():
        cursor.execute("INSERT INTO historial_sesion (hora, archivo, registros, confiabilidad) VALUES (?, ?, ?, ?)",
                       (hora, archivo, registros, confiabilidad))
        conn.commit()
    conn.close()

def cargar_historial_db():
    conn = _conectar_db()
    df_db = pd.read_sql_query("SELECT hora AS Hora, archivo AS Archivo, registros AS Registros, confiabilidad AS Confiabilidad FROM historial_sesion", conn)
    conn.close()
    return df_db.to_dict("records")

def guardar_cambio_db(fecha_hora, fila_indice, columna, valor_anterior, nuevo_valor, inspector):
    conn = _conectar_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO bitacora_cambios (fecha_hora, fila_indice, columna, valor_anterior, nuevo_valor, inspector) VALUES (?, ?, ?, ?, ?, ?)",
                   (fecha_hora, fila_indice, columna, valor_anterior, nuevo_valor, inspector))
    conn.commit()
    conn.close()

def cargar_bitacora_db():
    conn = _conectar_db()
    df_bit = pd.read_sql_query("SELECT fecha_hora AS 'Fecha/Hora', fila_indice AS 'Fila Índice', columna AS 'Columna Modificada', valor_anterior AS 'Valor Anterior', nuevo_valor AS 'Nuevo Valor', inspector AS 'Inspector' FROM bitacora_cambios", conn)
    conn.close()
    return df_bit.to_dict("records")

def guardar_contenedor_db(booking, contenedor, p_linea, p_senasa, destino, estado):
    inicializar_base_datos()
    conn = _conectar_db()
    cursor = conn.cursor()
    cursor.execute(
        """
        INSERT INTO contenedores_despacho
            (booking, contenedor, precinto_linea, precinto_senasa, destino, estado)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (booking, contenedor, p_linea, p_senasa, destino, estado),
    )
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()
    return new_id


def cargar_contenedores_db(limite: int = 100):
    inicializar_base_datos()
    conn = _conectar_db()
    df_cont = pd.read_sql_query(
        """
        SELECT booking AS 'Booking', contenedor AS 'Contenedor',
               precinto_linea AS 'Precinto Línea', precinto_senasa AS 'Precinto SENASA',
               destino AS 'Destino', estado AS 'Estado'
        FROM contenedores_despacho
        ORDER BY rowid DESC
        LIMIT ?
        """,
        conn,
        params=(int(limite),),
    )
    conn.close()
    return df_cont.to_dict("records")


def buscar_contenedor_local(contenedor: str = "", booking: str = "") -> dict | None:
    """Busca un contenedor o booking en SQLite local."""
    inicializar_base_datos()
    c_id = str(contenedor or "").strip().upper()
    bkg = str(booking or "").strip().upper()
    if not c_id and not bkg:
        return None
    conn = _conectar_db()
    cur = conn.cursor()
    if c_id:
        cur.execute(
            """
            SELECT booking, contenedor, precinto_linea, precinto_senasa, destino, estado
            FROM contenedores_despacho
            WHERE upper(contenedor) = ?
            ORDER BY rowid DESC LIMIT 1
            """,
            (c_id,),
        )
        row = cur.fetchone()
        if row:
            conn.close()
            return dict(
                zip(
                    ["booking", "contenedor", "precinto_linea", "precinto_senasa", "destino", "estado"],
                    row,
                )
            )
    if bkg:
        cur.execute(
            """
            SELECT booking, contenedor, precinto_linea, precinto_senasa, destino, estado
            FROM contenedores_despacho
            WHERE upper(booking) = ?
            ORDER BY rowid DESC LIMIT 1
            """,
            (bkg,),
        )
        row = cur.fetchone()
        if row:
            conn.close()
            return dict(
                zip(
                    ["booking", "contenedor", "precinto_linea", "precinto_senasa", "destino", "estado"],
                    row,
                )
            )
    conn.close()
    return None


_RE_ISO6346 = re.compile(r"\b([A-Z]{4}\s?\d{7})\b", re.IGNORECASE)
_RE_BOOKING = re.compile(
    r"(?:booking|bkg|bl|reserva)\s*[:=]\s*([A-Za-z0-9\-/]+)", re.IGNORECASE
)
_RE_CONT_KV = re.compile(
    r"(?:contenedor|container|cntr|reefer)\s*[:=]\s*([A-Za-z0-9]+)", re.IGNORECASE
)
_RE_PREC_LIN = re.compile(
    r"(?:precinto_linea|precinto.?linea|seal.?line|seal)\s*[:=]\s*([A-Za-z0-9\-/]+)",
    re.IGNORECASE,
)
_RE_PREC_SEN = re.compile(
    r"(?:precinto_senasa|senasa|precinto.?oficial)\s*[:=]\s*([A-Za-z0-9\-/]+)",
    re.IGNORECASE,
)


def parsear_payload_contenedor(texto: str) -> dict:
    """
    Extrae booking / contenedor / precintos del QR o texto.
    Formatos: JSON, clave=valor, pipe, código ISO 6346 puro.
    """
    crudo = (texto or "").strip()
    out = {
        "texto": crudo,
        "booking": None,
        "contenedor": None,
        "precinto_linea": None,
        "precinto_senasa": None,
        "destino": None,
        "estado": None,
    }
    if not crudo:
        return out

    if crudo.startswith("{") and crudo.endswith("}"):
        try:
            obj = json.loads(crudo)
            if isinstance(obj, dict):
                for k, alias in (
                    ("booking", ("booking", "bkg", "bl", "reserva")),
                    ("contenedor", ("contenedor", "container", "cntr", "reefer", "id")),
                    ("precinto_linea", ("precinto_linea", "precinto_line", "seal", "seal_line")),
                    ("precinto_senasa", ("precinto_senasa", "senasa", "precinto_oficial")),
                    ("destino", ("destino", "market", "mercado")),
                    ("estado", ("estado", "status")),
                ):
                    for a in alias:
                        if obj.get(a):
                            out[k] = str(obj[a]).strip()
                            break
                if out["contenedor"] or out["booking"]:
                    if out["contenedor"]:
                        out["contenedor"] = out["contenedor"].upper().replace(" ", "")
                    return out
        except json.JSONDecodeError:
            pass

    if "|" in crudo:
        partes = [p.strip() for p in crudo.split("|") if p.strip()]
        # booking|contenedor|p_linea|p_senasa|destino
        if len(partes) >= 2:
            out["booking"] = partes[0]
            out["contenedor"] = partes[1].upper().replace(" ", "")
            if len(partes) >= 3:
                out["precinto_linea"] = partes[2]
            if len(partes) >= 4:
                out["precinto_senasa"] = partes[3]
            if len(partes) >= 5:
                out["destino"] = partes[4]
            return out

    m_iso = _RE_ISO6346.search(crudo.upper().replace(" ", ""))
    if m_iso or _RE_ISO6346.search(crudo.upper()):
        m2 = _RE_ISO6346.search(crudo.upper())
        if m2:
            out["contenedor"] = m2.group(1).replace(" ", "")

    m_b = _RE_BOOKING.search(crudo)
    if m_b:
        out["booking"] = m_b.group(1).strip()
    m_c = _RE_CONT_KV.search(crudo)
    if m_c:
        out["contenedor"] = m_c.group(1).strip().upper().replace(" ", "")
    m_pl = _RE_PREC_LIN.search(crudo)
    if m_pl:
        out["precinto_linea"] = m_pl.group(1).strip()
    m_ps = _RE_PREC_SEN.search(crudo)
    if m_ps:
        out["precinto_senasa"] = m_ps.group(1).strip()

    # Texto corto = posible contenedor ISO
    if not out["contenedor"] and re.fullmatch(r"[A-Za-z]{4}\d{7}", crudo.replace(" ", "")):
        out["contenedor"] = crudo.replace(" ", "").upper()
    elif not out["contenedor"] and not out["booking"] and len(crudo) < 40:
        # asumir booking corto
        out["booking"] = crudo

    return out


def consultar_contenedores_supabase(
    contenedor: str | None = None,
    booking: str | None = None,
    timeout: float = 15.0,
) -> dict:
    """
    GET REST ≈ supabase.table('contenedores_despacho').select('*').eq(...)
    """
    url, key, err_cfg = _supabase_config()
    if err_cfg or not url or not key:
        return {
            "ok": False,
            "configurado": False,
            "encontrado": False,
            "filas": [],
            "mensaje": err_cfg or "Supabase no configurado",
            "endpoint": None,
        }

    params = {"select": "*", "limit": "10"}
    c_id = (contenedor or "").strip()
    bkg = (booking or "").strip()
    if c_id:
        params["contenedor"] = f"eq.{c_id}"
    if bkg:
        params["booking"] = f"eq.{bkg}"
    if "contenedor" not in params and "booking" not in params:
        return {
            "ok": False,
            "configurado": True,
            "encontrado": False,
            "filas": [],
            "mensaje": "Se requiere contenedor o booking para consultar.",
            "endpoint": None,
        }

    endpoint = f"{url}/rest/v1/contenedores_despacho?{urllib.parse.urlencode(params)}"
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Accept": "application/json",
    }
    try:
        req = urllib.request.Request(endpoint, method="GET", headers=headers)
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            status = getattr(resp, "status", 200) or 200
            raw = resp.read().decode("utf-8", errors="replace")
            filas = []
            if raw.strip():
                data = json.loads(raw)
                if isinstance(data, list):
                    filas = data
                elif isinstance(data, dict):
                    filas = [data]
            return {
                "ok": True,
                "configurado": True,
                "encontrado": len(filas) > 0,
                "filas": filas,
                "status": status,
                "mensaje": f"Consulta OK: {len(filas)} fila(s).",
                "endpoint": endpoint,
            }
    except urllib.error.HTTPError as e:
        try:
            detalle = e.read().decode("utf-8", errors="replace")[:300]
        except Exception:
            detalle = str(e)
        return {
            "ok": False,
            "configurado": True,
            "encontrado": False,
            "filas": [],
            "status": e.code,
            "mensaje": f"HTTP {e.code}: {detalle}",
            "endpoint": endpoint,
        }
    except Exception as e:
        return {
            "ok": False,
            "configurado": True,
            "encontrado": False,
            "filas": [],
            "mensaje": f"{type(e).__name__}: {e}",
            "endpoint": endpoint,
        }


def enviar_contenedor_supabase(registro: dict, timeout: float = 15.0) -> dict:
    """POST a public.contenedores_despacho (variantes de columnas)."""
    url, key, err_cfg = _supabase_config()
    if err_cfg or not url or not key:
        return {
            "ok": False,
            "configurado": False,
            "mensaje": err_cfg or "Supabase no configurado",
            "payload": {},
        }

    endpoint = f"{url}/rest/v1/contenedores_despacho"
    base = {
        "booking": registro.get("booking"),
        "contenedor": registro.get("contenedor"),
        "precinto_linea": registro.get("precinto_linea"),
        "precinto_senasa": registro.get("precinto_senasa"),
        "destino": registro.get("destino"),
        "estado": registro.get("estado"),
        "inspector": registro.get("inspector"),
        "fecha": registro.get("fecha")
        or datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    base = {k: v for k, v in base.items() if v is not None and str(v).strip() != ""}

    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Prefer": "return=representation",
    }

    variantes = [
        base,
        {
            k: base[k]
            for k in ("booking", "contenedor", "precinto_linea", "precinto_senasa", "destino", "estado")
            if k in base
        },
        {k: base[k] for k in ("booking", "contenedor", "estado") if k in base},
    ]

    ultimo = "sin detalle"
    for pl in variantes:
        if not pl or "contenedor" not in pl:
            continue
        try:
            body = json.dumps(pl, ensure_ascii=False).encode("utf-8")
            req = urllib.request.Request(endpoint, data=body, method="POST", headers=headers)
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                status = getattr(resp, "status", 200) or 200
                if status in (200, 201, 204):
                    return {
                        "ok": True,
                        "configurado": True,
                        "status": status,
                        "mensaje": f"INSERT OK contenedores_despacho (HTTP {status})",
                        "payload": pl,
                        "endpoint": endpoint,
                    }
        except urllib.error.HTTPError as e:
            try:
                ultimo = f"HTTP {e.code}: {e.read().decode('utf-8', errors='replace')[:300]}"
            except Exception:
                ultimo = f"HTTP {e.code}"
            if e.code == 409:
                return {
                    "ok": True,
                    "ya_existia": True,
                    "configurado": True,
                    "status": 409,
                    "mensaje": "Contenedor ya registrado en Supabase (sin duplicar).",
                    "payload": pl,
                    "endpoint": endpoint,
                }
            continue
        except Exception as e:
            ultimo = f"{type(e).__name__}: {e}"
            continue

    return {
        "ok": False,
        "configurado": True,
        "mensaje": f"No se pudo insertar en Supabase contenedores_despacho: {ultimo}",
        "payload": base,
        "endpoint": endpoint,
    }


def validar_y_consultar_contenedor(
    imagen_bytes: bytes | None = None,
    texto_qr: str | None = None,
) -> dict:
    """
    Pipeline tipo Módulo 6: imagen/texto → QR → parse → local + Supabase.
    """
    textos = []
    decode_info = None
    if imagen_bytes:
        decode_info = decodificar_qr_desde_imagen(imagen_bytes)
        if not decode_info.get("ok"):
            return {
                "verificado": False,
                "tipo_ui": "error",
                "mensaje_ui": decode_info.get("mensaje") or "No se pudo leer el QR.",
                "payload": None,
                "local": None,
                "supabase": None,
                "decode": decode_info,
            }
        textos = list(decode_info.get("textos") or [])
    elif texto_qr and str(texto_qr).strip():
        textos = [str(texto_qr).strip()]
    else:
        return {
            "verificado": False,
            "tipo_ui": "error",
            "mensaje_ui": "Capture o pegue el QR / código del contenedor.",
            "payload": None,
            "local": None,
            "supabase": None,
            "decode": None,
        }

    payload = None
    for t in textos:
        p = parsear_payload_contenedor(t)
        if p.get("contenedor") or p.get("booking"):
            payload = p
            break
    if payload is None:
        payload = parsear_payload_contenedor(textos[0])

    c_id = payload.get("contenedor")
    bkg = payload.get("booking")
    if not c_id and not bkg:
        return {
            "verificado": False,
            "tipo_ui": "error",
            "mensaje_ui": (
                "🚨 No se detectó booking ni contenedor en el QR. "
                "Use JSON, ISO 6346 (4 letras + 7 dígitos) o booking|contenedor|…"
            ),
            "payload": payload,
            "local": None,
            "supabase": None,
            "decode": decode_info,
        }

    local = buscar_contenedor_local(contenedor=c_id or "", booking=bkg or "")
    sb = consultar_contenedores_supabase(contenedor=c_id, booking=bkg if not c_id else None)

    if local or (sb.get("ok") and sb.get("encontrado")):
        reg = local or ((sb.get("filas") or [{}])[0])
        return {
            "verificado": True,
            "tipo_ui": "success",
            "mensaje_ui": (
                f"✅ Contenedor verificado: `{reg.get('contenedor') or c_id or 'N/D'}` · "
                f"estado **{reg.get('estado') or 'REGISTRADO'}**"
            ),
            "payload": payload,
            "local": local,
            "supabase": sb,
            "registro": reg,
            "decode": decode_info,
        }

    return {
        "verificado": False,
        "tipo_ui": "error",
        "mensaje_ui": (
            f"🚨 Contenedor no encontrado en base local/nube. "
            f"Booking: {bkg or 'N/D'} · Contenedor: {c_id or 'N/D'}. "
            "Complete el formulario y selle el despacho."
        ),
        "payload": payload,
        "local": None,
        "supabase": sb,
        "registro": None,
        "decode": decode_info,
    }


def procesar_escaneo_contenedor_camara(imagen_bytes: bytes) -> dict:
    """Alias Módulo 5: cámara → consulta contenedor."""
    return validar_y_consultar_contenedor(imagen_bytes=imagen_bytes)


def registrar_contenedor_despacho(
    booking: str,
    contenedor: str,
    precinto_linea: str = "",
    precinto_senasa: str = "",
    destino: str = "",
    estado: str = "SELLADO Y LISTO",
    inspector: str = "",
) -> dict:
    """
    Guarda en SQLite y replica a Supabase (mismo patrón que frío / QR).
    """
    booking = str(booking or "").strip()
    contenedor = str(contenedor or "").strip().upper().replace(" ", "")
    precinto_linea = str(precinto_linea or "").strip()
    precinto_senasa = str(precinto_senasa or "").strip()
    destino = str(destino or "").strip() or "N/D"
    estado = str(estado or "SELLADO Y LISTO").strip()
    inspector = str(inspector or "N/D").strip()

    if not booking or not contenedor or not precinto_linea:
        return {
            "ok": False,
            "tipo_ui": "error",
            "mensaje_ui": "⚠️ Complete Booking, Contenedor y Precinto de línea (obligatorios).",
            "sqlite_ok": False,
            "supabase": None,
        }

    # Ya existe local
    ya = buscar_contenedor_local(contenedor=contenedor)
    if ya:
        return {
            "ok": True,
            "ya_existia": True,
            "tipo_ui": "success",
            "mensaje_ui": (
                f"✅ Contenedor `{contenedor}` ya estaba registrado "
                f"(estado: {ya.get('estado')})."
            ),
            "registro": ya,
            "sqlite_ok": True,
            "supabase": None,
        }

    try:
        row_id = guardar_contenedor_db(
            booking, contenedor, precinto_linea, precinto_senasa, destino, estado
        )
        sqlite_ok = True
        sqlite_msg = f"SQLite OK (id {row_id})"
    except Exception as e:
        row_id = None
        sqlite_ok = False
        sqlite_msg = f"SQLite error: {e}"
        return {
            "ok": False,
            "tipo_ui": "error",
            "mensaje_ui": f"🚨 No se pudo guardar localmente: {e}",
            "sqlite_ok": False,
            "sqlite_msg": sqlite_msg,
            "supabase": None,
        }

    registro = {
        "booking": booking,
        "contenedor": contenedor,
        "precinto_linea": precinto_linea,
        "precinto_senasa": precinto_senasa,
        "destino": destino,
        "estado": estado,
        "inspector": inspector,
        "fecha": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    try:
        sb = enviar_contenedor_supabase(registro)
    except Exception as e:
        sb = {"ok": False, "configurado": True, "mensaje": f"{type(e).__name__}: {e}"}

    return {
        "ok": True,
        "ya_existia": False,
        "tipo_ui": "success",
        "mensaje_ui": (
            f"✅ Contenedor `{contenedor}` sellado y registrado "
            f"(booking {booking})."
        ),
        "registro": registro,
        "id_local": row_id,
        "sqlite_ok": sqlite_ok,
        "sqlite_msg": sqlite_msg,
        "supabase": sb,
    }


def guardar_frio_db(camara, temperatura, hora_registro, estado, inspector, producto=""):
    """Persiste lectura de frío en SQLite local."""
    inicializar_base_datos()
    conn = _conectar_db()
    cursor = conn.cursor()
    # Columna producto (migración suave)
    try:
        cursor.execute("ALTER TABLE control_frio ADD COLUMN producto TEXT")
    except sqlite3.OperationalError:
        pass
    cursor.execute(
        """
        INSERT INTO control_frio (camara, temperatura, hora_registro, estado, inspector, producto)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (camara, temperatura, hora_registro, estado, inspector, producto or ""),
    )
    conn.commit()
    new_id = cursor.lastrowid
    conn.close()
    return new_id


def cargar_frio_db(limite=50):
    conn = _conectar_db()
    # producto puede no existir en DBs antiguas
    try:
        df_frio = pd.read_sql_query(
            """
            SELECT hora_registro AS 'Fecha/Hora', camara AS 'Cámara',
                   temperatura AS 'Temp °C', estado AS 'Estado',
                   COALESCE(inspector, '') AS 'Inspector',
                   COALESCE(producto, '') AS 'Fruta'
            FROM control_frio
            ORDER BY id DESC
            LIMIT ?
            """,
            conn,
            params=(limite,),
        )
    except Exception:
        df_frio = pd.read_sql_query(
            """
            SELECT hora_registro AS 'Fecha/Hora', camara AS 'Cámara',
                   temperatura AS 'Temp °C', estado AS 'Estado',
                   COALESCE(inspector, '') AS 'Inspector'
            FROM control_frio
            ORDER BY id DESC
            LIMIT ?
            """,
            conn,
            params=(limite,),
        )
    conn.close()
    return df_frio.to_dict("records")


# Rangos óptimos comerciales reales (agroexportación)
RANGOS_TEMP_FRUTO = {
    "Palta Hass": {"min": 4.0, "max": 6.0},
    "Arándano": {"min": -0.5, "max": 0.5},
    "Espárrago": {"min": 2.0, "max": 4.0},
    "Uva Red Globe": {"min": -1.0, "max": 0.0},
}


def obtener_rango_frio_fruta(producto, temp_min_override=None, temp_max_override=None):
    """
    Devuelve (t_min, t_max) según la fruta del menú.
    Para 'Personalizado' usa overrides del sidebar si vienen.
    """
    nombre = str(producto or "").strip()
    if nombre in RANGOS_TEMP_FRUTO:
        r = RANGOS_TEMP_FRUTO[nombre]
        return float(r["min"]), float(r["max"])

    t_min = float(temp_min_override) if temp_min_override is not None else 0.0
    if temp_max_override is not None:
        t_max = float(temp_max_override)
    else:
        t_max = t_min + 2.0
    if t_max < t_min:
        t_min, t_max = t_max, t_min
    return t_min, t_max


def validar_temperatura_fruta(temperatura, producto, temp_min_override=None, temp_max_override=None):
    """Indica si la temperatura está dentro del rango comercial de la fruta."""
    t_min, t_max = obtener_rango_frio_fruta(producto, temp_min_override, temp_max_override)
    temp = float(temperatura)
    en_rango = t_min <= temp <= t_max
    return {
        "en_rango": en_rango,
        "temp_min": t_min,
        "temp_max": t_max,
        "temperatura": temp,
        "producto": str(producto or "").strip() or "N/D",
    }


def enviar_control_frio_supabase(registro: dict, timeout: float = 15.0) -> dict:
    """
    POST a public.control_frio en Supabase (REST).
    Campos preferidos en minúsculas: fecha, camara, temperatura, estado,
    inspector, producto, temp_min, temp_max.
    """
    url, key, err_cfg = _supabase_config()
    if err_cfg or not url or not key:
        return {
            "ok": False,
            "configurado": False,
            "mensaje": err_cfg or "Supabase no configurado",
            "payload": {},
        }

    endpoint = f"{url}/rest/v1/control_frio"
    # Payload principal + alias comunes (fecha/hora_registro)
    payload = {
        "fecha": registro.get("fecha") or registro.get("hora_registro"),
        "hora_registro": registro.get("hora_registro") or registro.get("fecha"),
        "camara": registro.get("camara"),
        "temperatura": registro.get("temperatura"),
        "estado": registro.get("estado"),
        "inspector": registro.get("inspector"),
        "producto": registro.get("producto"),
        "temp_min": registro.get("temp_min"),
        "temp_max": registro.get("temp_max"),
    }
    # Quitar Nones
    payload = {k: v for k, v in payload.items() if v is not None}

    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Prefer": "return=representation",
    }

    variantes = [
        payload,
        # mínimo común
        {
            k: payload[k]
            for k in ("fecha", "camara", "temperatura", "estado", "inspector", "producto")
            if k in payload
        },
        {
            k: payload[k]
            for k in ("hora_registro", "camara", "temperatura", "estado", "inspector")
            if k in payload
        },
    ]

    ultimo_error = "sin detalle"
    for pl in variantes:
        if not pl:
            continue
        try:
            body = json.dumps(pl, ensure_ascii=False).encode("utf-8")
            req = urllib.request.Request(endpoint, data=body, method="POST", headers=headers)
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                status = getattr(resp, "status", 200) or 200
                raw = resp.read().decode("utf-8", errors="replace")[:300]
                if status in (200, 201, 204):
                    return {
                        "ok": True,
                        "configurado": True,
                        "status": status,
                        "mensaje": f"INSERT OK public.control_frio (HTTP {status})",
                        "payload": pl,
                        "respuesta": raw,
                    }
                ultimo_error = f"HTTP {status}: {raw}"
        except urllib.error.HTTPError as e:
            try:
                ultimo_error = f"HTTP {e.code}: {e.read().decode('utf-8', errors='replace')[:400]}"
            except Exception:
                ultimo_error = f"HTTP {e.code}: {e}"
            continue
        except Exception as e:
            ultimo_error = f"{type(e).__name__}: {e}"
            continue

    return {
        "ok": False,
        "configurado": True,
        "mensaje": f"No se pudo insertar en Supabase control_frio: {ultimo_error}",
        "payload": payload,
    }


def registrar_control_frio(
    camara,
    temperatura,
    producto,
    inspector,
    temp_min_override=None,
    temp_max_override=None,
    hora_registro=None,
):
    """
    Valida la temperatura según rangos comerciales de la fruta seleccionada,
    guarda en SQLite y replica a Supabase (tabla control_frio).

    Returns dict con en_rango, estado, mensajes UI y resultado remoto.
    """
    validacion = validar_temperatura_fruta(
        temperatura, producto, temp_min_override, temp_max_override
    )
    t_min = validacion["temp_min"]
    t_max = validacion["temp_max"]
    temp = validacion["temperatura"]
    fruta = validacion["producto"]
    en_rango = validacion["en_rango"]

    if en_rango:
        estado = "FRIO_OPTIMO"
        mensaje_ui = "✅ Temperatura validada"
        tipo_ui = "success"
    else:
        estado = "RUPTURA_CADENA_FRIO"
        mensaje_ui = f"🚨 ALERTA ROJA: Ruptura de cadena de frío para {fruta}"
        tipo_ui = "error"

    if not hora_registro:
        hora_registro = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 1) SQLite local (siempre)
    try:
        row_id = guardar_frio_db(
            camara=camara,
            temperatura=temp,
            hora_registro=hora_registro,
            estado=estado,
            inspector=inspector or "N/D",
            producto=fruta,
        )
        sqlite_ok = True
        sqlite_msg = f"SQLite OK (id {row_id})"
    except Exception as e:
        row_id = None
        sqlite_ok = False
        sqlite_msg = f"SQLite error: {e}"

    # 2) Supabase (siempre se intenta registrar, incluso en alerta)
    registro_remoto = {
        "fecha": hora_registro,
        "hora_registro": hora_registro,
        "camara": camara,
        "temperatura": temp,
        "estado": estado,
        "inspector": inspector or "N/D",
        "producto": fruta,
        "temp_min": t_min,
        "temp_max": t_max,
    }
    try:
        sb = enviar_control_frio_supabase(registro_remoto)
    except Exception as e:
        sb = {
            "ok": False,
            "configurado": True,
            "mensaje": f"{type(e).__name__}: {e}",
            "payload": registro_remoto,
        }

    # 3) Aviso WhatsApp / email solo en ruptura
    aviso = {"ok": False, "configurado": False, "mensaje": ""}
    if not en_rango:
        try:
            aviso = notificar_ruptura_frio(
                {
                    "camara": camara,
                    "temperatura": temp,
                    "temp_min": t_min,
                    "temp_max": t_max,
                    "producto": fruta,
                    "inspector": inspector or "N/D",
                    "hora_registro": hora_registro,
                    "estado": estado,
                }
            )
        except Exception as e:
            aviso = {
                "ok": False,
                "configurado": True,
                "mensaje": f"Aviso error: {type(e).__name__}: {e}",
            }

    return {
        "en_rango": en_rango,
        "estado": estado,
        "temp_min": t_min,
        "temp_max": t_max,
        "temperatura": temp,
        "producto": fruta,
        "camara": camara,
        "hora_registro": hora_registro,
        "tipo_ui": tipo_ui,
        "mensaje_ui": mensaje_ui,
        "sqlite_ok": sqlite_ok,
        "sqlite_msg": sqlite_msg,
        "id_local": row_id,
        "supabase": sb,
        "aviso": aviso,
    }

def cargar_datos_archivo(uploaded_file):
    if uploaded_file.name.endswith(".csv"):
        df = pd.read_csv(uploaded_file, dtype=str)
    else:
        df = pd.read_excel(uploaded_file, dtype=str)
    return df.map(lambda x: str(x).strip() if pd.notna(x) else "")

def resaltar_errores_celdas(val):
    val_str = str(val).strip()
    if val_str == "" or val_str == "-":
        return "background-color: #ffcccc; color: #990000; font-weight: bold;"
    return ""


# Paleta corporativa Excel
_EXCEL_AZUL_OSCURO = "1F4E78"
_EXCEL_AZUL_CLARO = "D6E3F0"
_EXCEL_ROJO_VACIO = "FFCCCC"
_EXCEL_BLANCO = "FFFFFF"
_EXCEL_TEXTO = "1A1A1A"


def _borde_tabla_excel():
    thin = Side(style="thin", color="B0B0B0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def aplicar_estilo_corporativo_hoja(worksheet, titulo_reporte: str, n_cols: int, n_filas_datos: int):
    """
    Aplica diseño corporativo a una hoja ya escrita con:
    - Fila 1: título del reporte (merge)
    - Fila 2: encabezados de columnas (azul oscuro, blanco, negrita)
    - Datos desde fila 3
    - Anchos automáticos y bordes limpios
    """
    if n_cols < 1:
        return

    azul_fill = PatternFill(start_color=_EXCEL_AZUL_OSCURO, end_color=_EXCEL_AZUL_OSCURO, fill_type="solid")
    azul_claro_fill = PatternFill(start_color=_EXCEL_AZUL_CLARO, end_color=_EXCEL_AZUL_CLARO, fill_type="solid")
    blanco_font = Font(name="Calibri", bold=True, color=_EXCEL_BLANCO, size=11)
    titulo_font = Font(name="Calibri", bold=True, color=_EXCEL_BLANCO, size=14)
    datos_font = Font(name="Calibri", color=_EXCEL_TEXTO, size=10)
    border = _borde_tabla_excel()
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- Encabezado elegante (título) ---
    if n_cols > 1:
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = worksheet.cell(row=1, column=1, value=titulo_reporte)
    title_cell.font = titulo_font
    title_cell.fill = azul_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 28
    # Pintar celdas merged del título
    for c in range(1, n_cols + 1):
        cell = worksheet.cell(row=1, column=c)
        cell.fill = azul_fill
        cell.border = border

    # --- Fila de títulos de columna (fila 2) ---
    worksheet.row_dimensions[2].height = 22
    for c in range(1, n_cols + 1):
        cell = worksheet.cell(row=2, column=c)
        cell.font = blanco_font
        cell.fill = azul_fill
        cell.alignment = align_center
        cell.border = border

    # --- Cuerpo de datos ---
    ultima_fila = 2 + max(n_filas_datos, 0)
    rojo_fill = PatternFill(start_color=_EXCEL_ROJO_VACIO, end_color=_EXCEL_ROJO_VACIO, fill_type="solid")
    for r in range(3, ultima_fila + 1):
        zebra = (r % 2 == 0)
        for c in range(1, n_cols + 1):
            cell = worksheet.cell(row=r, column=c)
            cell.font = datos_font
            cell.alignment = align_left
            cell.border = border
            val = cell.value
            if val is None or str(val).strip() in ("", "-"):
                cell.fill = rojo_fill
            elif zebra:
                cell.fill = azul_claro_fill

    # --- Ancho de columna automático (textos largos, stretch film, etc.) ---
    for c in range(1, n_cols + 1):
        col_letter = get_column_letter(c)
        max_len = 10
        for r in range(2, ultima_fila + 1):
            val = worksheet.cell(row=r, column=c).value
            if val is None:
                continue
            # Medir la línea más larga si hay saltos
            for parte in str(val).splitlines() or [""]:
                max_len = max(max_len, len(parte))
        # Margen + límite para no inflar a columnas absurdas
        ancho = min(max(max_len + 3, 12), 60)
        worksheet.column_dimensions[col_letter].width = ancho

    worksheet.freeze_panes = "A3"
    worksheet.print_title_rows = "1:2"


def escribir_dataframe_corporativo(workbook, sheet_name: str, df: pd.DataFrame, titulo: str):
    """Crea una hoja con título + headers corporativos + datos del DataFrame."""
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(title=sheet_name[:31])

    df_out = df.copy() if df is not None else pd.DataFrame()
    if df_out.empty and df_out.columns.empty:
        df_out = pd.DataFrame({"Aviso": ["Sin datos"]})

    n_cols = max(len(df_out.columns), 1)
    # Fila 1 reservada para título
    ws.cell(row=1, column=1, value="")
    # Fila 2: nombres de columna
    for c, col_name in enumerate(df_out.columns, start=1):
        ws.cell(row=2, column=c, value=str(col_name))
    # Filas de datos desde 3
    for r_idx, row in enumerate(df_out.itertuples(index=False), start=3):
        for c_idx, valor in enumerate(row, start=1):
            if pd.isna(valor):
                valor = ""
            ws.cell(row=r_idx, column=c_idx, value=valor)

    aplicar_estilo_corporativo_hoja(ws, titulo, n_cols, len(df_out))
    return ws


def generar_excel_corporativo(
    hojas: dict,
    titulo_general: str = "Reporte corporativo de exportación",
) -> io.BytesIO:
    """
    Genera un Excel multi-hoja con diseño corporativo.

    hojas: dict[str, tuple[pd.DataFrame, str]]  # sheet_name -> (df, titulo_hoja)
           o dict[str, pd.DataFrame]  # título = sheet_name
    """
    from openpyxl import Workbook

    wb = Workbook()
    # Eliminar hoja por defecto
    default = wb.active
    wb.remove(default)

    for nombre, contenido in hojas.items():
        if isinstance(contenido, tuple):
            df, titulo = contenido
        else:
            df, titulo = contenido, f"{titulo_general} — {nombre}"
        escribir_dataframe_corporativo(wb, nombre, df, titulo)

    if not wb.sheetnames:
        escribir_dataframe_corporativo(wb, "Reporte", pd.DataFrame({"Aviso": ["Sin datos"]}), titulo_general)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def extraer_sello_ecc_pdf(pdf_bytes) -> dict:
    """
    Extrae mensaje firmado, firma y llave pública desde un PDF ejecutivo.
    Busca marcadores [ECC_*] y también etiquetas legibles Firma:/Llave Pública:.
    """
    import re
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(pdf_bytes))
    texto = "\n".join((page.extract_text() or "") for page in reader.pages)

    def _marker(tag):
        m = re.search(rf"\[{tag}\](.*?)\[/{tag}\]", texto, flags=re.IGNORECASE | re.DOTALL)
        return m.group(1).strip() if m else ""

    mensaje = _marker("ECC_MSG")
    firma = _marker("ECC_SIG")
    publica = _marker("ECC_PUB")

    if not mensaje:
        m = re.search(r"Mensaje firmado:\s*(.+)", texto, flags=re.IGNORECASE)
        if m:
            mensaje = m.group(1).strip()
    if not firma:
        m = re.search(r"Firma:\s*([0-9a-fA-F]+)", texto)
        if m:
            firma = m.group(1).strip()
    if not publica:
        m = re.search(r"Llave P[uú]blica:\s*([0-9a-fA-F]+)", texto, flags=re.IGNORECASE)
        if m:
            publica = m.group(1).strip()

    return {
        "mensaje": mensaje,
        "firma": "".join(firma.split()),
        "llave_publica": "".join(publica.split()),
        "texto_extraido": texto,
    }


def generar_pdf_resumen(
    archivo_nombre,
    total_filas,
    total_errores,
    duplicados,
    confiabilidad,
    producto,
    auditor,
    congelado_estado,
    mercado,
    capa_texto,
    firma_ECDSA,
    llave_publica,
    mensaje_firmado="",
    planta_nombre="",
):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )
    story = []
    styles = getSampleStyleSheet()

    titulo_estilo = ParagraphStyle(
        'TituloReporte',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1F4E78'),
        spaceAfter=10
    )

    title = Paragraph("<b>REPORTE EJECUTIVO DE AUDITORÍA Y TRAZABILIDAD AGROINDUSTRIAL</b>", titulo_estilo)
    story.append(title)
    story.append(Spacer(1, 5))

    fecha_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    estado_txt = "APROBADO Y CONGELADO" if congelado_estado else "EN EDICIÓN / REVISIÓN"
    
    sub_planta = f"<b>Planta:</b> {planta_nombre} | " if str(planta_nombre).strip() else ""
    sub = Paragraph(
        f"{sub_planta}<b>Producto:</b> {producto} | <b>Destino:</b> {mercado}<br/>"
        f"<b>Estado:</b> {estado_txt} | <b>Fecha:</b> {fecha_str}",
        styles["Normal"],
    )
    story.append(sub)
    story.append(Spacer(1, 10))

    datos_tabla = [
        ["Indicador de Control", "Valor Registrado"],
        ["Nombre del Archivo Base", archivo_nombre],
        ["Total Registros Auditados", str(total_filas)],
        ["Celdas Vacías / Errores Base", str(total_errores)],
        ["Registros Duplicados", str(duplicados)],
        ["Índice de Confiabilidad Inicial", f"{confiabilidad}%"],
        ["Responsable de Auditoría", auditor],
    ]

    t = Table(datos_tabla, colWidths=[220, 280])
    t.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F2F2F2")),
            ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#D9D9D9")),
        ])
    )
    story.append(t)
    story.append(Spacer(1, 10))

    if capa_texto.strip() != "":
        capa_p = Paragraph(f"<b>Acción Correctiva (CAPA / Justificación):</b><br/>{capa_texto}", styles["Normal"])
        story.append(capa_p)
        story.append(Spacer(1, 10))

    mensaje_ref = mensaje_firmado or (
        f"Auditoría de Planta - Cultivo: {producto} - Fecha: {fecha_str[:10]} - Registros: {total_filas}"
    )
    # Marcadores machine-readable para verificación pública posterior
    bloque_verificacion = (
        f"[ECC_MSG]{mensaje_ref}[/ECC_MSG] "
        f"[ECC_SIG]{firma_ECDSA}[/ECC_SIG] "
        f"[ECC_PUB]{llave_publica}[/ECC_PUB] "
        f"[ECC_HASH]{calcular_hash_reporte(mensaje_ref, firma_ECDSA, llave_publica)}[/ECC_HASH]"
    )
    sello_data = [
        [Paragraph("<b>Sello Criptográfico Ed25519:</b>", styles["Normal"])],
        [Paragraph(f"<font size=7 face='Courier'><b>Mensaje firmado:</b> {mensaje_ref}</font>", styles["Normal"])],
        [Paragraph(f"<font size=7 face='Courier'><b>Firma:</b> {firma_ECDSA}</font>", styles["Normal"])],
        [Paragraph(f"<font size=7 face='Courier'><b>Llave Pública:</b> {llave_publica}</font>", styles["Normal"])],
        [Paragraph(f"<font size=5 face='Courier'>{bloque_verificacion}</font>", styles["Normal"])],
        [Paragraph(
            "<i>Verifique este PDF en la pestaña pública de la plataforma. "
            "Si el contenido firmado fue alterado, la verificación ECC fallará.</i>",
            styles["Normal"],
        )],
    ]
    
    tabla_sello = Table(sello_data, colWidths=[500])
    tabla_sello.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#EDF2F7')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#1F4E78')),
        ('PADDING', (0,0), (-1,-1), 8),
    ]))
    
    story.append(tabla_sello)
    doc.build(story)
    buffer.seek(0)
    return buffer

def generar_pdf_errores(df_errores, archivo_nombre, producto, auditor):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20,
    )
    story = []
    styles = getSampleStyleSheet()

    title = Paragraph("<b>REPORTE DE REGISTROS CON ERRORES O FALTANTES EN PLANTA</b>", styles["Title"])
    story.append(title)
    story.append(Spacer(1, 5))

    fecha_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sub = Paragraph(
        f"<b>Archivo Base:</b> {archivo_nombre} | <b>Cultivo:</b> {producto}<br/><b>Inspector:</b> {auditor} | <b>Fecha de Emisión:</b> {fecha_str}",
        styles["Normal"],
    )
    story.append(sub)
    story.append(Spacer(1, 10))

    if df_errores.empty:
        story.append(Paragraph("<b>¡Excelente! No se encontraron registros con errores o celdas vacías.</b>", styles["Normal"]))
    else:
        cols_a_mostrar = list(df_errores.columns[:7])
        tabla_datos = [[Paragraph(f"<b>{c}</b>", styles["Normal"]) for c in cols_a_mostrar]]
        
        for _, row in df_errores.iterrows():
            fila_cells = [Paragraph(str(row[c]), styles["Normal"]) for c in cols_a_mostrar]
            tabla_datos.append(fila_cells)

        ancho_col = 572 / len(cols_a_mostrar)
        t = Table(tabla_datos, colWidths=[ancho_col] * len(cols_a_mostrar))
        t.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C00000")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F9F9F9")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
            ])
        )
        story.append(t)

    doc.build(story)
    buffer.seek(0)
    return buffer


# ─── Módulo 7: Alertas y tendencias (inteligencia operativa de planta) ───────
# Fase A: reglas + estadísticas. Fase B: anomalías por z-score (sin sklearn).


def cargar_frio_dataframe(limite: int = 500) -> pd.DataFrame:
    """Devuelve lecturas de control_frio como DataFrame para análisis."""
    inicializar_base_datos()
    conn = _conectar_db()
    try:
        df = pd.read_sql_query(
            """
            SELECT id, camara, temperatura, hora_registro, estado,
                   COALESCE(inspector, '') AS inspector,
                   COALESCE(producto, '') AS producto
            FROM control_frio
            ORDER BY id DESC
            LIMIT ?
            """,
            conn,
            params=(int(limite),),
        )
    except Exception:
        df = pd.read_sql_query(
            """
            SELECT id, camara, temperatura, hora_registro, estado,
                   COALESCE(inspector, '') AS inspector
            FROM control_frio
            ORDER BY id DESC
            LIMIT ?
            """,
            conn,
            params=(int(limite),),
        )
        if not df.empty:
            df["producto"] = ""
    conn.close()
    if df.empty:
        return df
    df["temperatura"] = pd.to_numeric(df["temperatura"], errors="coerce")
    df["hora_dt"] = pd.to_datetime(df["hora_registro"], errors="coerce")
    return df


def _serie_zscore(valores: pd.Series) -> pd.Series:
    """Z-score robusto; NaN si no hay desviación usable."""
    s = pd.to_numeric(valores, errors="coerce")
    media = s.mean()
    std = s.std(ddof=0)
    if pd.isna(std) or std < 1e-9:
        return pd.Series([float("nan")] * len(s), index=s.index)
    return (s - media) / std


def _alerta(severidad: str, origen: str, titulo: str, detalle: str, metrica: str = "") -> dict:
    return {
        "severidad": severidad,  # critica | advertencia | info
        "origen": origen,
        "titulo": titulo,
        "detalle": detalle,
        "metrica": metrica,
    }


def analizar_tendencias_frio(
    limite: int = 500,
    z_umbral: float = 2.5,
    min_muestras_anomalia: int = 8,
) -> dict:
    """
    Analiza historial de cadena de frío (SQLite).
    Reglas: rupturas, % fuera de rango por cámara, tendencia reciente.
    Anomalías: |z-score| > umbral por cámara (cuando hay historial suficiente).
    """
    df = cargar_frio_dataframe(limite=limite)
    alertas: list[dict] = []
    por_camara: list[dict] = []
    serie_chart: pd.DataFrame | None = None

    if df.empty:
        return {
            "ok": True,
            "sin_datos": True,
            "total_lecturas": 0,
            "rupturas": 0,
            "pct_ruptura": 0.0,
            "alertas": [
                _alerta(
                    "info",
                    "frio",
                    "Sin historial de frío",
                    "Registre lecturas en el Módulo 4 para activar tendencias y anomalías.",
                )
            ],
            "por_camara": [],
            "serie_chart": None,
            "resumen": "Sin lecturas de cadena de frío en SQLite.",
        }

    total = len(df)
    es_ruptura = df["estado"].astype(str).str.upper().str.contains("RUPTURA", na=False)
    rupturas = int(es_ruptura.sum())
    pct_ruptura = round((rupturas / total) * 100, 1) if total else 0.0

    if pct_ruptura >= 20:
        alertas.append(
            _alerta(
                "critica",
                "frio",
                "Alta tasa de ruptura de frío",
                f"{rupturas} de {total} lecturas fuera de rango ({pct_ruptura}%).",
                f"{pct_ruptura}%",
            )
        )
    elif pct_ruptura >= 8:
        alertas.append(
            _alerta(
                "advertencia",
                "frio",
                "Rupturas de frío recurrentes",
                f"{rupturas} lecturas fuera de rango ({pct_ruptura}%). Revisar cámaras.",
                f"{pct_ruptura}%",
            )
        )

    # Serie para gráfico (cronológica)
    df_ord = df.dropna(subset=["temperatura"]).sort_values("hora_dt", ascending=True)
    if not df_ord.empty:
        serie_chart = (
            df_ord.assign(etiqueta=df_ord["hora_registro"].astype(str).str[-8:])
            .groupby(["camara", "etiqueta"], as_index=False)["temperatura"]
            .mean()
        )

    for camara, g in df.groupby(df["camara"].astype(str)):
        g = g.dropna(subset=["temperatura"]).copy()
        n = len(g)
        if n == 0:
            continue
        rup_c = int(g["estado"].astype(str).str.upper().str.contains("RUPTURA", na=False).sum())
        pct_c = round((rup_c / n) * 100, 1)
        media = round(float(g["temperatura"].mean()), 2)
        std = round(float(g["temperatura"].std(ddof=0) or 0), 2)

        # Tendencia: últimas 5 vs anteriores (usa id si la hora empata)
        sort_cols = [c for c in ("hora_dt", "id") if c in g.columns]
        g_ord = g.sort_values(sort_cols, ascending=True) if sort_cols else g
        tendencia = "estable"
        delta = 0.0
        if n >= 6:
            ult = g_ord["temperatura"].tail(5).mean()
            prev = g_ord["temperatura"].iloc[:-5]
            ant = prev.tail(5).mean() if len(prev) else ult
            delta = round(float(ult - ant), 2)
            if delta >= 0.8:
                tendencia = "calentamiento"
            elif delta <= -0.8:
                tendencia = "enfriamiento"

        anomalias = 0
        if n >= min_muestras_anomalia:
            z = _serie_zscore(g_ord["temperatura"])
            anomalias = int((z.abs() > z_umbral).sum())

        fila = {
            "camara": camara,
            "lecturas": n,
            "rupturas": rup_c,
            "pct_ruptura": pct_c,
            "temp_media": media,
            "temp_std": std,
            "tendencia": tendencia,
            "delta_reciente": delta,
            "anomalias_z": anomalias,
        }
        por_camara.append(fila)

        if pct_c >= 25 and n >= 4:
            alertas.append(
                _alerta(
                    "critica",
                    "frio",
                    f"Cámara crítica: {camara}",
                    f"{rup_c}/{n} rupturas ({pct_c}%). Media {media} °C.",
                    f"{pct_c}%",
                )
            )
        elif tendencia == "calentamiento" and delta >= 0.8:
            alertas.append(
                _alerta(
                    "advertencia",
                    "frio",
                    f"Tendencia al alza: {camara}",
                    f"Últimas lecturas +{delta} °C vs. tramo anterior. Anticipar revisión.",
                    f"+{delta} °C",
                )
            )
        if anomalias > 0:
            alertas.append(
                _alerta(
                    "advertencia",
                    "frio",
                    f"Anomalía térmica: {camara}",
                    f"{anomalias} lectura(s) con |z-score| > {z_umbral} (patrón atípico).",
                    f"{anomalias} outlier(s)",
                )
            )

    # Detalle de rupturas recientes (máx. 3; evita ruido si la tasa ya es alta)
    if 0 < rupturas <= 8:
        recientes = df.loc[es_ruptura].head(3)
        for _, row in recientes.iterrows():
            alertas.append(
                _alerta(
                    "advertencia",
                    "frio",
                    "Ruptura registrada",
                    f"{row.get('camara')} · {row.get('temperatura')} °C · "
                    f"{row.get('hora_registro')} · {row.get('producto') or 'N/D'}",
                    str(row.get("estado") or ""),
                )
            )

    n_crit = sum(1 for a in alertas if a["severidad"] == "critica")
    n_adv = sum(1 for a in alertas if a["severidad"] == "advertencia")
    resumen = (
        f"{total} lecturas · {rupturas} rupturas ({pct_ruptura}%) · "
        f"{n_crit} crítica(s) · {n_adv} advertencia(s)"
    )

    return {
        "ok": True,
        "sin_datos": False,
        "total_lecturas": total,
        "rupturas": rupturas,
        "pct_ruptura": pct_ruptura,
        "alertas": alertas,
        "por_camara": sorted(por_camara, key=lambda x: (-x["pct_ruptura"], -x["anomalias_z"])),
        "serie_chart": serie_chart,
        "resumen": resumen,
    }


def analizar_patrones_packing(
    df: pd.DataFrame,
    cols_mapa: dict | None = None,
    peso_min_caja: float = 4.0,
    max_merma_permitida: float = 5.0,
    z_umbral_peso: float = 2.5,
) -> dict:
    """
    Patrones del archivo de packing cargado: pesos, LMR por productor/fundo, merma.
    """
    alertas: list[dict] = []
    cols_mapa = cols_mapa or mapear_columnas_trazabilidad(df)
    total = len(df) if df is not None else 0
    if df is None or total == 0:
        return {
            "ok": True,
            "sin_datos": True,
            "alertas": [
                _alerta("info", "packing", "Sin archivo de packing", "Cargue un Excel/CSV para analizar patrones.")
            ],
            "por_productor": [],
            "stats_peso": {},
            "resumen": "Sin datos de packing.",
        }

    # Pesos
    col_peso = cols_mapa.get("peso")
    stats_peso: dict = {}
    if col_peso and col_peso in df.columns:
        pesos = pd.to_numeric(df[col_peso], errors="coerce")
        validos = pesos.dropna()
        bajo_min = int((pesos < float(peso_min_caja)).fillna(False).sum())
        media = round(float(validos.mean()), 2) if len(validos) else 0.0
        std = round(float(validos.std(ddof=0) or 0), 2) if len(validos) else 0.0
        stats_peso = {
            "media": media,
            "std": std,
            "bajo_minimo": bajo_min,
            "n_validos": int(len(validos)),
            "vacios": int(pesos.isna().sum() + (df[col_peso].astype(str).str.strip() == "").sum()),
        }
        if bajo_min > 0:
            sev = "critica" if bajo_min >= max(3, int(total * 0.05)) else "advertencia"
            alertas.append(
                _alerta(
                    sev,
                    "packing",
                    "Cajas bajo peso mínimo",
                    f"{bajo_min} registro(s) < {peso_min_caja} kg (media lote {media} kg).",
                    f"{bajo_min} cajas",
                )
            )
        if len(validos) >= 10 and std > 0:
            z = _serie_zscore(validos)
            outliers = int((z.abs() > z_umbral_peso).sum())
            stats_peso["outliers_z"] = outliers
            if outliers > 0:
                alertas.append(
                    _alerta(
                        "advertencia",
                        "packing",
                        "Pesos atípicos (z-score)",
                        f"{outliers} caja(s) con peso fuera del patrón del lote (|z| > {z_umbral_peso}).",
                        f"{outliers} outlier(s)",
                    )
                )

    # LMR por productor / fundo
    col_lmr = cols_mapa.get("lmr")
    col_prod = cols_mapa.get("productor")
    col_fundo = cols_mapa.get("fundo")
    por_productor: list[dict] = []

    if col_lmr and col_lmr in df.columns:
        estados = df[col_lmr].astype(str).map(interpretar_estado_lmr)
        n_rech = int((estados == "rechazado").sum())
        n_alerta = int((estados == "alerta").sum())
        if n_rech > 0:
            alertas.append(
                _alerta(
                    "critica",
                    "packing",
                    "LMR rechazado en archivo",
                    f"{n_rech} registro(s) con veredicto de rechazo / supera LMR.",
                    f"{n_rech}",
                )
            )
        elif n_alerta > 0:
            alertas.append(
                _alerta(
                    "advertencia",
                    "packing",
                    "LMR en zona de alerta",
                    f"{n_alerta} registro(s) cercanos al límite / cuarentena.",
                    f"{n_alerta}",
                )
            )

        grupo_col = col_prod if col_prod and col_prod in df.columns else None
        if grupo_col:
            tmp = df.copy()
            tmp["_estado_lmr"] = estados
            tmp["_grupo"] = tmp[grupo_col].astype(str).str.strip().replace("", "N/D")
            for nombre, g in tmp.groupby("_grupo"):
                n = len(g)
                rech = int((g["_estado_lmr"] == "rechazado").sum())
                aler = int((g["_estado_lmr"] == "alerta").sum())
                riesgo = round(((rech + aler) / n) * 100, 1) if n else 0.0
                fila = {
                    "productor": nombre,
                    "registros": n,
                    "lmr_alerta": aler,
                    "lmr_rechazo": rech,
                    "pct_riesgo": riesgo,
                }
                if col_fundo and col_fundo in g.columns:
                    fundos = sorted({str(x).strip() for x in g[col_fundo] if str(x).strip()})
                    fila["fundos"] = ", ".join(fundos[:4]) + ("…" if len(fundos) > 4 else "")
                else:
                    fila["fundos"] = "N/D"
                por_productor.append(fila)
                if riesgo >= 30 and n >= 3:
                    alertas.append(
                        _alerta(
                            "critica" if rech > 0 else "advertencia",
                            "packing",
                            f"Productor con patrón LMR: {nombre}",
                            f"{rech} rechazo(s) y {aler} alerta(s) en {n} cajas ({riesgo}%).",
                            f"{riesgo}%",
                        )
                    )
            por_productor.sort(key=lambda x: (-x["pct_riesgo"], -x["lmr_rechazo"]))

    # Merma / descarte
    col_cat = None
    for c in df.columns:
        cu = str(c).upper()
        if "CATEGORIA" in cu or cu in ("CAT", "CATEGORÍA"):
            col_cat = c
            break
    pct_merma = 0.0
    if col_cat:
        cats = df[col_cat].astype(str).str.upper().str.strip()
        descarte = int(cats.isin(["DESCARTE", "MERMA", "RECHAZO"]).sum())
        pct_merma = round((descarte / total) * 100, 2) if total else 0.0
        if pct_merma > float(max_merma_permitida):
            alertas.append(
                _alerta(
                    "critica",
                    "packing",
                    "Merma sobre el límite",
                    f"Merma/descarte {pct_merma}% (límite {max_merma_permitida}%).",
                    f"{pct_merma}%",
                )
            )

    vacios = int((df.astype(str).apply(lambda s: s.str.strip()) == "").sum().sum())
    if total and vacios > total * 2:
        alertas.append(
            _alerta(
                "advertencia",
                "packing",
                "Alta incompleción de datos",
                f"{vacios} celdas vacías en el lote. Revisar antes de congelar.",
                f"{vacios} vacíos",
            )
        )

    n_crit = sum(1 for a in alertas if a["severidad"] == "critica")
    n_adv = sum(1 for a in alertas if a["severidad"] == "advertencia")
    resumen = f"{total} filas · merma {pct_merma}% · {n_crit} crítica(s) · {n_adv} advertencia(s)"

    return {
        "ok": True,
        "sin_datos": False,
        "alertas": alertas,
        "por_productor": por_productor,
        "stats_peso": stats_peso,
        "pct_merma": pct_merma,
        "total_filas": total,
        "resumen": resumen,
    }


def consolidar_inteligencia_planta(
    df_packing: pd.DataFrame | None = None,
    cols_mapa: dict | None = None,
    peso_min_caja: float = 4.0,
    max_merma_permitida: float = 5.0,
    limite_frio: int = 500,
) -> dict:
    """
    Combina tendencias de frío + patrones de packing en un informe único.
    """
    frio = analizar_tendencias_frio(limite=limite_frio)
    packing = (
        analizar_patrones_packing(
            df_packing,
            cols_mapa=cols_mapa,
            peso_min_caja=peso_min_caja,
            max_merma_permitida=max_merma_permitida,
        )
        if df_packing is not None
        else {
            "ok": True,
            "sin_datos": True,
            "alertas": [],
            "por_productor": [],
            "stats_peso": {},
            "resumen": "Packing no cargado en esta sesión.",
        }
    )

    alertas = list(frio.get("alertas") or []) + list(packing.get("alertas") or [])
    orden = {"critica": 0, "advertencia": 1, "info": 2}
    alertas.sort(key=lambda a: (orden.get(a.get("severidad"), 9), a.get("origen", "")))

    n_crit = sum(1 for a in alertas if a["severidad"] == "critica")
    n_adv = sum(1 for a in alertas if a["severidad"] == "advertencia")
    n_info = sum(1 for a in alertas if a["severidad"] == "info")

    if n_crit:
        veredicto = "ACCION_REQUERIDA"
        veredicto_ui = "Se detectaron alertas críticas. Revisar antes de cerrar el lote o despachar."
    elif n_adv:
        veredicto = "VIGILANCIA"
        veredicto_ui = "Hay tendencias a vigilar. Anticipar revisión de frío / calidad."
    else:
        veredicto = "ESTABLE"
        veredicto_ui = "Sin patrones de riesgo relevantes en los datos disponibles."

    return {
        "ok": True,
        "veredicto": veredicto,
        "veredicto_ui": veredicto_ui,
        "contadores": {"criticas": n_crit, "advertencias": n_adv, "info": n_info},
        "alertas": alertas,
        "frio": frio,
        "packing": packing,
    }


# ─── Dashboard de turno + alertas de frío activas ─────────────────────────────


def _parse_fecha_hora_flexible(valor) -> datetime.datetime | None:
    if valor is None:
        return None
    s = str(valor).strip()
    if not s:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%H:%M:%S",
    ):
        try:
            dt = datetime.datetime.strptime(s, fmt)
            if fmt == "%H:%M:%S":
                hoy = datetime.date.today()
                return datetime.datetime.combine(hoy, dt.time())
            return dt
        except ValueError:
            continue
    try:
        return pd.to_datetime(s, errors="coerce").to_pydatetime()
    except Exception:
        return None


def alertas_frio_activas(limite: int = 30, horas_ventana: int = 12) -> dict:
    """
    Rupturas de cadena de frío recientes que requieren atención del turno.
    """
    inicializar_base_datos()
    df = cargar_frio_dataframe(limite=max(limite * 3, 100))
    if df.empty:
        return {
            "ok": True,
            "activas": [],
            "total": 0,
            "nivel": "ESTABLE",
            "mensaje": "Sin lecturas de frío registradas.",
        }

    es_ruptura = df["estado"].astype(str).str.upper().str.contains("RUPTURA", na=False)
    rupturas = df.loc[es_ruptura].copy()
    if rupturas.empty:
        return {
            "ok": True,
            "activas": [],
            "total": 0,
            "nivel": "ESTABLE",
            "mensaje": "Sin rupturas de frío en el historial reciente.",
        }

    ahora = datetime.datetime.now()
    corte = ahora - datetime.timedelta(hours=int(horas_ventana))
    activas = []
    for _, row in rupturas.iterrows():
        dt = _parse_fecha_hora_flexible(row.get("hora_registro"))
        if dt is not None and dt < corte:
            continue
        activas.append(
            {
                "camara": str(row.get("camara") or "N/D"),
                "temperatura": row.get("temperatura"),
                "hora": str(row.get("hora_registro") or ""),
                "producto": str(row.get("producto") or ""),
                "inspector": str(row.get("inspector") or ""),
                "estado": str(row.get("estado") or ""),
            }
        )
        if len(activas) >= limite:
            break

    total = len(activas)
    if total >= 3:
        nivel = "CRITICO"
        mensaje = f"{total} rupturas de frío en las últimas {horas_ventana} h."
    elif total >= 1:
        nivel = "ALERTA"
        mensaje = f"{total} ruptura(s) de frío reciente(s)."
    else:
        nivel = "ESTABLE"
        mensaje = f"Sin rupturas en las últimas {horas_ventana} h."

    return {
        "ok": True,
        "activas": activas,
        "total": total,
        "nivel": nivel,
        "mensaje": mensaje,
        "horas_ventana": horas_ventana,
    }


def resumen_dashboard_turno(fecha: datetime.date | None = None) -> dict:
    """
    KPIs del turno / día para el jefe de planta:
    sellos ECC, rupturas de frío, contenedores, cargas de packing.
    """
    inicializar_base_datos()
    dia = fecha or datetime.date.today()
    dia_str = dia.strftime("%Y-%m-%d")
    conn = _conectar_db()

    def _count_like(sql: str, params=()):
        try:
            cur = conn.execute(sql, params)
            row = cur.fetchone()
            return int(row[0] if row and row[0] is not None else 0)
        except Exception:
            return 0

    # historial_reportes: fecha_hora suele ser 'YYYY-MM-DD HH:MM:SS'
    sellos_hoy = _count_like(
        "SELECT COUNT(*) FROM historial_reportes WHERE substr(fecha_hora, 1, 10) = ?",
        (dia_str,),
    )
    # control_frio
    lecturas_hoy = _count_like(
        "SELECT COUNT(*) FROM control_frio WHERE substr(hora_registro, 1, 10) = ?",
        (dia_str,),
    )
    rupturas_hoy = _count_like(
        """
        SELECT COUNT(*) FROM control_frio
        WHERE substr(hora_registro, 1, 10) = ?
          AND upper(COALESCE(estado, '')) LIKE '%RUPTURA%'
        """,
        (dia_str,),
    )
    # Si hora_registro solo trae HH:MM:SS, contar por id reciente del día vía fallback
    if lecturas_hoy == 0:
        try:
            df_f = pd.read_sql_query(
                "SELECT hora_registro, estado FROM control_frio ORDER BY id DESC LIMIT 200",
                conn,
            )
            n_lec = 0
            n_rup = 0
            for _, r in df_f.iterrows():
                dt = _parse_fecha_hora_flexible(r.get("hora_registro"))
                if dt is None or dt.date() != dia:
                    continue
                n_lec += 1
                if "RUPTURA" in str(r.get("estado") or "").upper():
                    n_rup += 1
            lecturas_hoy = n_lec
            rupturas_hoy = n_rup
        except Exception:
            pass

    contenedores_hoy = _count_like("SELECT COUNT(*) FROM contenedores_despacho")
    # No hay fecha en contenedores: mostrar total reciente (últimos IDs del día no aplica)
    try:
        contenedores_total = _count_like("SELECT COUNT(*) FROM contenedores_despacho")
    except Exception:
        contenedores_total = 0

    cargas_packing = _count_like("SELECT COUNT(*) FROM historial_sesion")

    lotes_sellados = []
    try:
        df_s = pd.read_sql_query(
            """
            SELECT fecha_hora, lote, responsable, producto, registros
            FROM historial_reportes
            WHERE substr(fecha_hora, 1, 10) = ?
            ORDER BY id DESC
            LIMIT 15
            """,
            conn,
            params=(dia_str,),
        )
        lotes_sellados = df_s.to_dict("records") if not df_s.empty else []
    except Exception:
        lotes_sellados = []

    conn.close()

    frio = alertas_frio_activas(limite=10, horas_ventana=12)

    if frio["nivel"] == "CRITICO" or rupturas_hoy >= 3:
        estado_turno = "CRITICO"
    elif frio["nivel"] == "ALERTA" or rupturas_hoy >= 1:
        estado_turno = "VIGILANCIA"
    else:
        estado_turno = "ESTABLE"

    return {
        "ok": True,
        "fecha": dia_str,
        "estado_turno": estado_turno,
        "kpis": {
            "sellos_ecc": sellos_hoy,
            "lecturas_frio": lecturas_hoy,
            "rupturas_frio": rupturas_hoy,
            "contenedores": contenedores_total,
            "cargas_packing": cargas_packing,
        },
        "lotes_sellados": lotes_sellados,
        "alertas_frio": frio,
    }


def generar_pdf_dashboard_turno(
    dash: dict,
    planta_nombre: str = "",
    inspector: str = "",
    rol: str = "",
):
    """PDF ejecutivo del dashboard de turno (KPIs + alertas de frío)."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )
    story = []
    styles = getSampleStyleSheet()
    titulo_estilo = ParagraphStyle(
        "TituloDash",
        parent=styles["Heading1"],
        fontSize=15,
        textColor=colors.HexColor("#1F4E78"),
        spaceAfter=8,
    )
    story.append(Paragraph("<b>DASHBOARD DE TURNO — RESUMEN OPERATIVO</b>", titulo_estilo))

    fecha_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fecha_turno = dash.get("fecha") or fecha_str[:10]
    estado = dash.get("estado_turno") or "—"
    sub_bits = []
    if str(planta_nombre).strip():
        sub_bits.append(f"<b>Planta:</b> {planta_nombre}")
    sub_bits.append(f"<b>Turno:</b> {fecha_turno}")
    sub_bits.append(f"<b>Estado:</b> {estado}")
    if str(inspector).strip():
        sub_bits.append(f"<b>Emitido por:</b> {inspector}")
    if str(rol).strip():
        sub_bits.append(f"<b>Rol:</b> {rol}")
    sub_bits.append(f"<b>Generado:</b> {fecha_str}")
    story.append(Paragraph(" | ".join(sub_bits), styles["Normal"]))
    story.append(Spacer(1, 12))

    k = dash.get("kpis") or {}
    datos_kpi = [
        ["Indicador", "Valor"],
        ["Sellos ECC hoy", str(k.get("sellos_ecc", 0))],
        ["Lecturas de frío hoy", str(k.get("lecturas_frio", 0))],
        ["Rupturas de frío hoy", str(k.get("rupturas_frio", 0))],
        ["Contenedores (total)", str(k.get("contenedores", 0))],
        ["Cargas packing (sesión)", str(k.get("cargas_packing", 0))],
    ]
    t = Table(datos_kpi, colWidths=[260, 220])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F2F2F2")),
                ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#D9D9D9")),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 14))

    af = dash.get("alertas_frio") or {}
    story.append(
        Paragraph(
            f"<b>Alertas de frío:</b> {af.get('mensaje') or 'Sin detalle'}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 6))
    activas = af.get("activas") or []
    if not activas:
        story.append(Paragraph("Sin rupturas activas en la ventana de alerta.", styles["Normal"]))
    else:
        filas_af = [["Cámara", "Temp.", "Producto", "Inspector", "Hora", "Estado"]]
        for a in activas[:20]:
            filas_af.append(
                [
                    str(a.get("camara") or "—")[:24],
                    str(a.get("temperatura") if a.get("temperatura") is not None else "—"),
                    str(a.get("producto") or "—")[:18],
                    str(a.get("inspector") or "—")[:18],
                    str(a.get("hora") or "—")[:19],
                    str(a.get("estado") or "—")[:22],
                ]
            )
        ta = Table(filas_af, colWidths=[75, 45, 70, 70, 85, 85])
        ta.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8B2942")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FFF5F5")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(ta)

    story.append(Spacer(1, 14))
    lotes = dash.get("lotes_sellados") or []
    story.append(Paragraph(f"<b>Sellos ECC del día ({len(lotes)}):</b>", styles["Normal"]))
    story.append(Spacer(1, 4))
    if not lotes:
        story.append(Paragraph("Aún no hay sellos archivados hoy.", styles["Normal"]))
    else:
        filas_l = [["Hora", "Lote", "Responsable", "Producto", "Registros"]]
        for L in lotes[:15]:
            filas_l.append(
                [
                    str(L.get("fecha_hora") or "—")[:19],
                    str(L.get("lote") or "—")[:20],
                    str(L.get("responsable") or "—")[:18],
                    str(L.get("producto") or "—")[:16],
                    str(L.get("registros") if L.get("registros") is not None else "—"),
                ]
            )
        tl = Table(filas_l, colWidths=[95, 90, 90, 85, 60])
        tl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F2F2F2")),
                ]
            )
        )
        story.append(tl)

    story.append(Spacer(1, 16))
    story.append(
        Paragraph(
            "<i>Documento operativo de turno. No sustituye el reporte ejecutivo con sello Ed25519.</i>",
            styles["Normal"],
        )
    )
    doc.build(story)
    buffer.seek(0)
    return buffer


# ─── Avisos por correo / WhatsApp (ruptura de frío) ───────────────────────────


def _leer_secret_aviso(*rutas):
    """Lee un valor de secrets por varias rutas posibles. Retorna str o ''."""
    try:
        import streamlit as st

        secrets = st.secrets
    except Exception:
        return ""

    def _get(obj, key):
        try:
            return obj[key]
        except Exception:
            return None

    for ruta in rutas:
        cur = secrets
        ok = True
        for parte in ruta:
            cur = _get(cur, parte)
            if cur is None:
                ok = False
                break
        if ok and str(cur).strip():
            return str(cur).strip()
    return ""


def _config_avisos() -> dict:
    """
    Configuración opcional en secrets [avisos]:
    email_to, smtp_host, smtp_port, smtp_user, smtp_pass, smtp_from,
    whatsapp_to, callmebot_apikey, twilio_sid, twilio_token, twilio_from
    """
    cfg = {
        "email_to": _leer_secret_aviso(
            ("avisos", "email_to"), ("credenciales", "AVISO_EMAIL_TO"), ("AVISO_EMAIL_TO",)
        ),
        "smtp_host": _leer_secret_aviso(
            ("avisos", "smtp_host"), ("credenciales", "SMTP_HOST"), ("SMTP_HOST",)
        )
        or "smtp.gmail.com",
        "smtp_port": _leer_secret_aviso(
            ("avisos", "smtp_port"), ("credenciales", "SMTP_PORT"), ("SMTP_PORT",)
        )
        or "587",
        "smtp_user": _leer_secret_aviso(
            ("avisos", "smtp_user"), ("credenciales", "SMTP_USER"), ("SMTP_USER",)
        ),
        "smtp_pass": _leer_secret_aviso(
            ("avisos", "smtp_pass"), ("credenciales", "SMTP_PASS"), ("SMTP_PASS",)
        ),
        "smtp_from": _leer_secret_aviso(
            ("avisos", "smtp_from"), ("credenciales", "SMTP_FROM"), ("SMTP_FROM",)
        ),
        "whatsapp_to": _leer_secret_aviso(
            ("avisos", "whatsapp_to"), ("credenciales", "WHATSAPP_TO"), ("WHATSAPP_TO",)
        ),
        "callmebot_apikey": _leer_secret_aviso(
            ("avisos", "callmebot_apikey"),
            ("credenciales", "CALLMEBOT_APIKEY"),
            ("CALLMEBOT_APIKEY",),
        ),
        "twilio_sid": _leer_secret_aviso(("avisos", "twilio_sid"), ("credenciales", "TWILIO_SID")),
        "twilio_token": _leer_secret_aviso(
            ("avisos", "twilio_token"), ("credenciales", "TWILIO_TOKEN")
        ),
        "twilio_from": _leer_secret_aviso(
            ("avisos", "twilio_from"), ("credenciales", "TWILIO_FROM")
        ),
    }
    if not cfg["smtp_from"]:
        cfg["smtp_from"] = cfg["smtp_user"]
    cfg["email_ok"] = bool(cfg["email_to"] and cfg["smtp_user"] and cfg["smtp_pass"])
    cfg["wa_callmebot_ok"] = bool(cfg["whatsapp_to"] and cfg["callmebot_apikey"])
    cfg["wa_twilio_ok"] = bool(
        cfg["whatsapp_to"] and cfg["twilio_sid"] and cfg["twilio_token"] and cfg["twilio_from"]
    )
    cfg["whatsapp_ok"] = cfg["wa_callmebot_ok"] or cfg["wa_twilio_ok"]
    cfg["habilitado"] = cfg["email_ok"] or cfg["whatsapp_ok"]
    return cfg


def _cooldown_aviso_ok(clave: str, minutos: int = 30) -> bool:
    """Evita spam: un aviso por cámara cada N minutos (SQLite)."""
    inicializar_base_datos()
    conn = _conectar_db()
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS avisos_enviados (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            clave TEXT NOT NULL,
            canal TEXT,
            fecha_hora TEXT NOT NULL,
            detalle TEXT
        )
        """
    )
    conn.commit()
    cur.execute(
        """
        SELECT fecha_hora FROM avisos_enviados
        WHERE clave = ?
        ORDER BY id DESC LIMIT 1
        """,
        (clave,),
    )
    row = cur.fetchone()
    ahora = datetime.datetime.now()
    if row:
        try:
            ult = datetime.datetime.strptime(str(row[0])[:19], "%Y-%m-%d %H:%M:%S")
            if (ahora - ult).total_seconds() < minutos * 60:
                conn.close()
                return False
        except Exception:
            pass
    conn.close()
    return True


def _registrar_aviso_enviado(clave: str, canal: str, detalle: str = ""):
    inicializar_base_datos()
    conn = _conectar_db()
    conn.execute(
        "INSERT INTO avisos_enviados (clave, canal, fecha_hora, detalle) VALUES (?, ?, ?, ?)",
        (
            clave,
            canal,
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            (detalle or "")[:400],
        ),
    )
    conn.commit()
    conn.close()


def enviar_aviso_email(asunto: str, cuerpo: str) -> dict:
    cfg = _config_avisos()
    if not cfg["email_ok"]:
        return {
            "ok": False,
            "configurado": False,
            "mensaje": "Email no configurado en secrets [avisos]",
        }
    try:
        import smtplib
        from email.mime.text import MIMEText

        msg = MIMEText(cuerpo, "plain", "utf-8")
        msg["Subject"] = asunto
        msg["From"] = cfg["smtp_from"]
        msg["To"] = cfg["email_to"]
        port = int(cfg["smtp_port"] or 587)
        with smtplib.SMTP(cfg["smtp_host"], port, timeout=20) as server:
            server.starttls()
            server.login(cfg["smtp_user"], cfg["smtp_pass"])
            server.sendmail(cfg["smtp_from"], [cfg["email_to"]], msg.as_string())
        return {
            "ok": True,
            "configurado": True,
            "mensaje": f"Email enviado a {cfg['email_to']}",
        }
    except Exception as e:
        return {
            "ok": False,
            "configurado": True,
            "mensaje": f"Email error: {type(e).__name__}: {e}",
        }


def enviar_aviso_whatsapp(texto: str) -> dict:
    cfg = _config_avisos()
    if not cfg["whatsapp_ok"]:
        return {
            "ok": False,
            "configurado": False,
            "mensaje": "WhatsApp no configurado (CallMeBot o Twilio en secrets [avisos])",
        }

    if cfg["wa_twilio_ok"]:
        try:
            import base64

            to = cfg["whatsapp_to"]
            if not to.startswith("whatsapp:"):
                to = "whatsapp:+" + to.lstrip("+")
            endpoint = (
                f"https://api.twilio.com/2010-04-01/Accounts/{cfg['twilio_sid']}/Messages.json"
            )
            data = urllib.parse.urlencode(
                {"From": cfg["twilio_from"], "To": to, "Body": texto[:1500]}
            ).encode("utf-8")
            req = urllib.request.Request(endpoint, data=data, method="POST")
            token = base64.b64encode(
                f"{cfg['twilio_sid']}:{cfg['twilio_token']}".encode()
            ).decode()
            req.add_header("Authorization", f"Basic {token}")
            req.add_header("Content-Type", "application/x-www-form-urlencoded")
            with urllib.request.urlopen(req, timeout=20) as resp:
                status = getattr(resp, "status", 200) or 200
                if status in (200, 201):
                    return {
                        "ok": True,
                        "configurado": True,
                        "mensaje": "WhatsApp Twilio OK",
                        "canal": "twilio",
                    }
                return {
                    "ok": False,
                    "configurado": True,
                    "mensaje": f"Twilio HTTP {status}",
                }
        except Exception as e:
            return {
                "ok": False,
                "configurado": True,
                "mensaje": f"Twilio error: {type(e).__name__}: {e}",
            }

    try:
        phone = cfg["whatsapp_to"].lstrip("+")
        q = urllib.parse.urlencode(
            {"phone": phone, "text": texto[:1000], "apikey": cfg["callmebot_apikey"]}
        )
        url = f"https://api.callmebot.com/whatsapp.php?{q}"
        req = urllib.request.Request(url, method="GET")
        with urllib.request.urlopen(req, timeout=25) as resp:
            raw = resp.read().decode("utf-8", errors="replace")[:200]
            status = getattr(resp, "status", 200) or 200
            if status == 200:
                return {
                    "ok": True,
                    "configurado": True,
                    "mensaje": f"WhatsApp CallMeBot OK",
                    "canal": "callmebot",
                    "detalle": raw,
                }
            return {
                "ok": False,
                "configurado": True,
                "mensaje": f"CallMeBot HTTP {status}: {raw}",
            }
    except Exception as e:
        return {
            "ok": False,
            "configurado": True,
            "mensaje": f"WhatsApp error: {type(e).__name__}: {e}",
        }


def notificar_ruptura_frio(registro: dict, forzar: bool = False) -> dict:
    """
    Envía aviso por email y/o WhatsApp si hay secrets.
    Cooldown 30 min por cámara (salvo forzar=True).
    """
    cfg = _config_avisos()
    if not cfg["habilitado"]:
        return {
            "ok": False,
            "configurado": False,
            "email": {"ok": False, "mensaje": "no config"},
            "whatsapp": {"ok": False, "mensaje": "no config"},
            "mensaje": "Avisos no configurados (secrets [avisos]).",
        }

    camara = str(registro.get("camara") or "N/D")
    clave = f"frio:{camara}"
    if not forzar and not _cooldown_aviso_ok(clave, minutos=30):
        return {
            "ok": False,
            "configurado": True,
            "omitido": True,
            "mensaje": f"Aviso omitido (cooldown 30 min · {camara})",
            "email": {},
            "whatsapp": {},
        }

    asunto = f"ALERTA frío · {camara}"
    cuerpo = (
        f"RUPTURA DE CADENA DE FRÍO\n"
        f"Cámara: {camara}\n"
        f"Temp: {registro.get('temperatura')} °C "
        f"(rango {registro.get('temp_min')}–{registro.get('temp_max')} °C)\n"
        f"Producto: {registro.get('producto')}\n"
        f"Inspector: {registro.get('inspector')}\n"
        f"Hora: {registro.get('hora_registro')}\n"
        f"Estado: {registro.get('estado')}\n"
        f"— Validador de planta"
    )

    email_r = (
        enviar_aviso_email(asunto, cuerpo)
        if cfg["email_ok"]
        else {"ok": False, "configurado": False, "mensaje": "Email no configurado"}
    )
    wa_r = (
        enviar_aviso_whatsapp(cuerpo)
        if cfg["whatsapp_ok"]
        else {"ok": False, "configurado": False, "mensaje": "WhatsApp no configurado"}
    )

    enviados = []
    if email_r.get("ok"):
        enviados.append("email")
        _registrar_aviso_enviado(clave, "email", email_r.get("mensaje", ""))
    if wa_r.get("ok"):
        enviados.append("whatsapp")
        _registrar_aviso_enviado(clave, "whatsapp", wa_r.get("mensaje", ""))

    ok = bool(enviados)
    return {
        "ok": ok,
        "configurado": True,
        "canales": enviados,
        "mensaje": (
            f"Aviso enviado: {', '.join(enviados)}"
            if ok
            else f"No se pudo enviar · email: {email_r.get('mensaje')} · wa: {wa_r.get('mensaje')}"
        ),
        "email": email_r,
        "whatsapp": wa_r,
    }
