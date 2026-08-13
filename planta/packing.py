"""Packing list: columnas, carga de archivo y Excel corporativo."""
from __future__ import annotations

import datetime
import hashlib
import io
import json
import os
import re
import sqlite3
import urllib.error
import urllib.parse
import urllib.request

import pandas as pd

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from planta.db import DB_NAME  # noqa: F401
ALIAS_COLUMNAS = {
    "id_unidad": ["CAJA", "SSCC", "PALLET", "CODIGO", "BARCODE", "EAN", "GTIN", "ID"],
    "lote": ["LOTE"],
    "fundo": ["FUNDO", "PARCELA", "CAMPO", "ORIGEN", "SECTOR"],
    "productor": ["PRODUCTOR", "PROVEEDOR", "AGRICULTOR"],
    "turno": ["TURNO", "LINEA", "PROCESO", "PACKING"],
    "peso": ["PESO"],
    "cosecha": ["COSECHA", "HARVEST", "FECHA_COSECHA", "F_COSECHA"],
    "lmr": ["LMR", "RESIDUO", "ANALISIS", "FITOSANITARIO"],
    "calibre": ["CALIBRE"],
    "categoria": ["CATEGORIA", "CAT"],
}


def encontrar_columna(df, keywords):
    """Devuelve el nombre real de la primera columna cuyo nombre contenga alguno de los keywords."""
    for col in df.columns:
        col_up = str(col).upper()
        for kw in keywords:
            if kw.upper() in col_up:
                return col
    return None


def mapear_columnas_trazabilidad(df):
    return {clave: encontrar_columna(df, aliases) for clave, aliases in ALIAS_COLUMNAS.items()}


CAMPOS_MAPEO_UI = (
    "id_unidad",
    "lote",
    "peso",
    "calibre",
    "fundo",
    "productor",
    "turno",
    "cosecha",
    "lmr",
    "categoria",
)
CAMPOS_CRITICOS_PACKING = ("lote", "peso", "calibre")


def resolver_mapa_columnas(df, mapeo_manual: dict | None = None) -> dict:
    """
    Combina detección automática por alias con un mapeo manual del usuario.
    En mapeo_manual: valor None o '' deja el auto; '(ninguna)' fuerza sin columna.
    """
    auto = mapear_columnas_trazabilidad(df)
    if not mapeo_manual:
        return auto
    out = dict(auto)
    cols = set(str(c) for c in df.columns)
    for clave, elegido in mapeo_manual.items():
        if clave not in out:
            continue
        if elegido is None or elegido == "" or elegido == "(auto)":
            continue
        if elegido == "(ninguna)":
            out[clave] = None
            continue
        if str(elegido) in cols:
            out[clave] = str(elegido)
    return out


def campos_criticos_sin_mapear(mapa: dict | None) -> list[str]:
    """Nombres legibles de LOTE/PESO/CALIBRE aún sin columna asignada."""
    mapa = mapa or {}
    return [k.upper() for k in CAMPOS_CRITICOS_PACKING if not mapa.get(k)]


def _valor_celda(fila, col, default="N/D"):
    if col is None or col not in fila.index:
        return default
    val = fila[col]
    if pd.isna(val) or str(val).strip() in ("", "-"):
        return default
    return str(val).strip()


def buscar_registros_por_codigo(df, codigo):
    """Busca filas cuyo ID (caja/pallet/sscc/código) o cualquier celda coincida con el código."""
    codigo = str(codigo).strip()
    if not codigo or df.empty:
        return df.iloc[0:0].copy()

    cols_mapa = mapear_columnas_trazabilidad(df)
    cols_prioridad = [
        c for c in [
            cols_mapa["id_unidad"],
            cols_mapa["lote"],
            encontrar_columna(df, ["CONTENEDOR", "BOOKING"]),
        ] if c is not None
    ]
    # Evitar duplicados manteniendo orden
    cols_busqueda = list(dict.fromkeys(cols_prioridad + list(df.columns)))

    mask = pd.Series(False, index=df.index)
    for col in cols_busqueda:
        serie = df[col].astype(str).str.strip()
        mask = mask | serie.str.fullmatch(codigo, case=False, na=False)
        if mask.any():
            break

    if not mask.any():
        # Búsqueda parcial solo en columnas de identificación
        for col in cols_prioridad:
            serie = df[col].astype(str).str.strip()
            mask = mask | serie.str.contains(codigo, case=False, na=False, regex=False)

    return df.loc[mask].copy()


def armar_arbol_trazabilidad(fila, cols_mapa, inspector, codigo_buscado):
    """Construye un diccionario legible del árbol genealógico a partir de una fila real."""
    return {
        "codigo": codigo_buscado,
        "fundo": _valor_celda(fila, cols_mapa["fundo"], "No informado en archivo"),
        "productor": _valor_celda(fila, cols_mapa["productor"], "No informado en archivo"),
        "lote": _valor_celda(fila, cols_mapa["lote"]),
        "turno": _valor_celda(fila, cols_mapa["turno"], "No informado en archivo"),
        "cosecha": _valor_celda(fila, cols_mapa["cosecha"], "No informado en archivo"),
        "peso": _valor_celda(fila, cols_mapa["peso"]),
        "calibre": _valor_celda(fila, cols_mapa["calibre"]),
        "lmr": _valor_celda(fila, cols_mapa["lmr"], "Sin dato LMR en archivo"),
        "inspector": inspector,
        "fila_indice": int(fila.name) if fila.name is not None else None,
    }


def interpretar_estado_lmr(texto):
    """Clasifica un resultado LMR textual en conforme / alerta / rechazado."""
    t = str(texto).strip().upper()
    if not t or t in ("N/D", "-", "SIN DATO LMR EN ARCHIVO", "NAN"):
        return "sin_dato"
    if any(k in t for k in ["RECHAZ", "SUPERA", "FAIL", "NO CONFORME", "BLOQUE"]):
        return "rechazado"
    if any(k in t for k in ["ALERTA", "CERCANO", "CUARENTENA", "PENDIENTE", "WARNING"]):
        return "alerta"
    if any(k in t for k in ["CONFORME", "APROB", "OK", "BAJO", "PASS", "CUMPLE"]):
        return "conforme"
    return "sin_dato"


def registrar_peso_ultima_fila(df, peso):
    """Inyecta el peso capturado en la última fila de la columna PESO (si existe)."""
    df_out = df.copy()
    col_peso = encontrar_columna(df_out, ALIAS_COLUMNAS["peso"])
    if col_peso is None or df_out.empty:
        return df_out, False, col_peso
    df_out.iloc[-1, df_out.columns.get_loc(col_peso)] = str(peso)
    return df_out, True, col_peso


# Tabla destino en Supabase: public.historial_reportes
SUPABASE_TABLA_SELLOS = "historial_reportes"
# Solo estas columnas (minúsculas exactas, sin extras)
SUPABASE_CAMPOS_SELLO = ("fecha", "lote", "hash_sha256", "inspector")


def cargar_datos_archivo(uploaded_file):
    if uploaded_file.name.endswith(".csv"):
        df = pd.read_csv(uploaded_file, dtype=str)
    else:
        df = pd.read_excel(uploaded_file, dtype=str)
    return df.map(lambda x: str(x).strip() if pd.notna(x) else "")

def resaltar_errores_celdas(val):
    val_str = str(val).strip()
    if val_str == "" or val_str == "-":
        return "background-color: #ffcccc; color: #990000; font-weight: bold;"
    return ""


# Paleta corporativa Excel
_EXCEL_AZUL_OSCURO = "1F4E78"
_EXCEL_AZUL_CLARO = "D6E3F0"
_EXCEL_ROJO_VACIO = "FFCCCC"
_EXCEL_BLANCO = "FFFFFF"
_EXCEL_TEXTO = "1A1A1A"


def _borde_tabla_excel():
    thin = Side(style="thin", color="B0B0B0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def aplicar_estilo_corporativo_hoja(worksheet, titulo_reporte: str, n_cols: int, n_filas_datos: int):
    """
    Aplica diseño corporativo a una hoja ya escrita con:
    - Fila 1: título del reporte (merge)
    - Fila 2: encabezados de columnas (azul oscuro, blanco, negrita)
    - Datos desde fila 3
    - Anchos automáticos y bordes limpios
    """
    if n_cols < 1:
        return

    azul_fill = PatternFill(start_color=_EXCEL_AZUL_OSCURO, end_color=_EXCEL_AZUL_OSCURO, fill_type="solid")
    azul_claro_fill = PatternFill(start_color=_EXCEL_AZUL_CLARO, end_color=_EXCEL_AZUL_CLARO, fill_type="solid")
    blanco_font = Font(name="Calibri", bold=True, color=_EXCEL_BLANCO, size=11)
    titulo_font = Font(name="Calibri", bold=True, color=_EXCEL_BLANCO, size=14)
    datos_font = Font(name="Calibri", color=_EXCEL_TEXTO, size=10)
    border = _borde_tabla_excel()
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- Encabezado elegante (título) ---
    if n_cols > 1:
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = worksheet.cell(row=1, column=1, value=titulo_reporte)
    title_cell.font = titulo_font
    title_cell.fill = azul_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[1].height = 28
    # Pintar celdas merged del título
    for c in range(1, n_cols + 1):
        cell = worksheet.cell(row=1, column=c)
        cell.fill = azul_fill
        cell.border = border

    # --- Fila de títulos de columna (fila 2) ---
    worksheet.row_dimensions[2].height = 22
    for c in range(1, n_cols + 1):
        cell = worksheet.cell(row=2, column=c)
        cell.font = blanco_font
        cell.fill = azul_fill
        cell.alignment = align_center
        cell.border = border

    # --- Cuerpo de datos ---
    ultima_fila = 2 + max(n_filas_datos, 0)
    rojo_fill = PatternFill(start_color=_EXCEL_ROJO_VACIO, end_color=_EXCEL_ROJO_VACIO, fill_type="solid")
    for r in range(3, ultima_fila + 1):
        zebra = (r % 2 == 0)
        for c in range(1, n_cols + 1):
            cell = worksheet.cell(row=r, column=c)
            cell.font = datos_font
            cell.alignment = align_left
            cell.border = border
            val = cell.value
            if val is None or str(val).strip() in ("", "-"):
                cell.fill = rojo_fill
            elif zebra:
                cell.fill = azul_claro_fill

    # --- Ancho de columna automático (textos largos, stretch film, etc.) ---
    for c in range(1, n_cols + 1):
        col_letter = get_column_letter(c)
        max_len = 10
        for r in range(2, ultima_fila + 1):
            val = worksheet.cell(row=r, column=c).value
            if val is None:
                continue
            # Medir la línea más larga si hay saltos
            for parte in str(val).splitlines() or [""]:
                max_len = max(max_len, len(parte))
        # Margen + límite para no inflar a columnas absurdas
        ancho = min(max(max_len + 3, 12), 60)
        worksheet.column_dimensions[col_letter].width = ancho

    worksheet.freeze_panes = "A3"
    worksheet.print_title_rows = "1:2"


def escribir_dataframe_corporativo(workbook, sheet_name: str, df: pd.DataFrame, titulo: str):
    """Crea una hoja con título + headers corporativos + datos del DataFrame."""
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    ws = workbook.create_sheet(title=sheet_name[:31])

    df_out = df.copy() if df is not None else pd.DataFrame()
    if df_out.empty and df_out.columns.empty:
        df_out = pd.DataFrame({"Aviso": ["Sin datos"]})

    n_cols = max(len(df_out.columns), 1)
    # Fila 1 reservada para título
    ws.cell(row=1, column=1, value="")
    # Fila 2: nombres de columna
    for c, col_name in enumerate(df_out.columns, start=1):
        ws.cell(row=2, column=c, value=str(col_name))
    # Filas de datos desde 3
    for r_idx, row in enumerate(df_out.itertuples(index=False), start=3):
        for c_idx, valor in enumerate(row, start=1):
            if pd.isna(valor):
                valor = ""
            ws.cell(row=r_idx, column=c_idx, value=valor)

    aplicar_estilo_corporativo_hoja(ws, titulo, n_cols, len(df_out))
    return ws


def generar_excel_corporativo(
    hojas: dict,
    titulo_general: str = "Reporte corporativo de exportación",
) -> io.BytesIO:
    """
    Genera un Excel multi-hoja con diseño corporativo.

    hojas: dict[str, tuple[pd.DataFrame, str]]  # sheet_name -> (df, titulo_hoja)
           o dict[str, pd.DataFrame]  # título = sheet_name
    """
    from openpyxl import Workbook

    wb = Workbook()
    # Eliminar hoja por defecto
    default = wb.active
    wb.remove(default)

    for nombre, contenido in hojas.items():
        if isinstance(contenido, tuple):
            df, titulo = contenido
        else:
            df, titulo = contenido, f"{titulo_general} — {nombre}"
        escribir_dataframe_corporativo(wb, nombre, df, titulo)

    if not wb.sheetnames:
        escribir_dataframe_corporativo(wb, "Reporte", pd.DataFrame({"Aviso": ["Sin datos"]}), titulo_general)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
