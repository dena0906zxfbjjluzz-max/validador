import streamlit as st
import datetime
import io
import pandas as pd

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

import funciones
import motor_planta

# --- CONFIGURACIÓN DE SEGURIDAD GENERAL ---
USUARIO_CORRECTO = "calidad"
PASSWORD_CORRECTO = "control2026"

if "autenticado" not in st.session_state:
    st.session_state["autenticado"] = False

if not st.session_state["autenticado"]:
    st.title("🔒 Acceso Restringido - Control de Calidad")
    st.write("Por favor, introduzca sus credenciales para ingresar a la plataforma corporativa.")
    
    usuarioInput = st.text_input("Usuario:")
    passwordInput = st.text_input("Contraseña:", type="password")
    
    if st.button("Ingresar"):
        if usuarioInput == USUARIO_CORRECTO and passwordInput == PASSWORD_CORRECTO:
            st.session_state["autenticado"] = True
            st.success("Acceso concedido.")
            st.rerun()
        else:
            st.error("Usuario o contraseña incorrectos. Inténtelo de nuevo.")
    st.stop()  
# -------------------------------------------

st.set_page_config(
    page_title="Sistema Integral de Exportación - Cerro Prieto",
    page_icon="📊",
    layout="wide",
)

st.markdown(
    """
    <style>
    .stButton>button {
        background-color: #1F4E78;
        color: white;
        border-radius: 8px;
        font-weight: bold;
        border: none;
        padding: 0.5rem 1rem;
    }
    .stButton>button:hover {
        background-color: #153859;
        color: white;
    }
    </style>
""",
    unsafe_allow_html=True,
)

st.title("🏭 Plataforma Corporativa de Control de Calidad, Trazabilidad, Contenedores y Planta (Perú)")
st.write(
    "Sistema integral optimizado con motor Rust nativo: GS1, Balanzas, LMR SENASA, Trazabilidad Inversa, Control de Frío y Gestión de Precintos de Contenedor."
)

funciones.inicializar_base_datos()

if "lote_congelado" not in st.session_state:
    st.session_state["lote_congelado"] = False

if "mostrar_vacios" not in st.session_state:
    st.session_state["mostrar_vacios"] = False

st.sidebar.header("⚙️ Parámetros de Planta & Destino")

auditor_nombre = st.sidebar.text_input("Inspector Asignado:", value="Control de Calidad")
producto_sel = st.sidebar.selectbox(
    "Cultivo / Fruta:", ["Palta Hass", "Arándano", "Espárrago", "Uva Red Globe", "Personalizado"]
)

mercado_destino = st.sidebar.selectbox(
    "🌍 Mercado de Destino:", ["Europa (GlobalGAP/UE)", "Estados Unidos (FDA/USDA)", "Asia / China", "Chile / Local"]
)

if producto_sel == "Palta Hass":
    limit_temp_default = 5.0
    limit_brix_default = 21.0
elif producto_sel == "Arándano":
    limit_temp_default = 1.5
    limit_brix_default = 11.5
elif producto_sel == "Espárrago":
    limit_temp_default = 2.0
    limit_brix_default = 8.0
elif producto_sel == "Uva Red Globe":
    limit_temp_default = 0.5
    limit_brix_default = 16.0
else:
    limit_temp_default = 4.0
    limit_brix_default = 10.0

temp_min_limite = st.sidebar.number_input("Temp. Mínima Cámara (°C):", value=limit_temp_default, step=0.5)
brix_min_limite = st.sidebar.number_input("Materia Seca / Brix Mínimo:", value=limit_brix_default, step=0.5)

st.sidebar.subheader("📦 Tolerancias Logísticas")
peso_min_caja = st.sidebar.number_input("Peso Mínimo Neto Caja (kg):", value=4.0, step=0.1)
max_merma_permitida = st.sidebar.number_input("Límite Máximo de Merma (%):", value=5.0, step=0.5)

archivo = st.file_uploader(
    "Cargar Base de Datos de Recepción / Empaque (Excel o CSV)", type=["xlsx", "csv"]
)

if archivo is not None:
    try:
        archivo_key = f"{archivo.name}_{archivo.size}"
        if st.session_state.get("archivo_activo") != archivo_key:
            st.session_state["archivo_activo"] = archivo_key
            st.session_state["df_trabajo"] = funciones.cargar_datos_archivo(archivo)
            st.session_state["lote_congelado"] = False

        df_original = st.session_state["df_trabajo"]
        cols_traza = funciones.mapear_columnas_trazabilidad(df_original)

        columnas_requeridas = ["LOTE", "PESO", "CALIBRE"]
        columnas_actuales_mayus = [c.upper() for c in df_original.columns]
        columnas_faltantes = [req for req in columnas_requeridas if not any(req in col for col in columnas_actuales_mayus)]

        if len(columnas_faltantes) > 0:
            st.warning(f"⚠️ **Aviso de Estructura:** El archivo no contiene columnas con nombres exactos como **{', '.join(columnas_faltantes)}**. La app operará con normalidad en modo flexible.")

        if "Observaciones_Rechazo" not in df_original.columns:
            df_original["Observaciones_Rechazo"] = ""
            st.session_state["df_trabajo"] = df_original

        total_filas = len(df_original)
        total_columnas = len(df_original.columns)
        total_duplicados = int(df_original.duplicated().sum())

        vacios_por_columna = (df_original == "").sum()
        total_errores = int(vacios_por_columna.sum())

        celdas_totales = total_filas * total_columnas
        
        try:
            eficiencia_rust, estado_rust = motor_planta.validar_datos_planta(
                float(total_filas), float(total_errores)
            )
            porcentaje_limpio = round(eficiencia_rust, 1)
            estado_rust = f"{estado_rust} [{motor_planta.motor_activo()}]"
        except Exception:
            porcentaje_limpio = (
                round(((celdas_totales - total_errores) / celdas_totales) * 100, 1)
                if celdas_totales > 0
                else 100
            )
            estado_rust = "Motor Python Respaldo"

        hora_actual = datetime.datetime.now().strftime("%H:%M:%S")
        funciones.guardar_historial_db(hora_actual, archivo.name, total_filas, f"{porcentaje_limpio}%")

        st.sidebar.markdown("---")
        st.sidebar.metric(label="Total Registros", value=total_filas)
        st.sidebar.metric(
            label="Campos Vacíos",
            value=total_errores,
            delta=f"-{total_errores}" if total_errores > 0 else "0",
            delta_color="inverse",
        )
        st.sidebar.metric(label="Confiabilidad (Motor Rust)", value=f"{porcentaje_limpio}%", delta=estado_rust)

        historial_db_data = funciones.cargar_historial_db()
        if historial_db_data:
            st.sidebar.subheader("🕒 Historial Persistente (SQLite)")
            st.sidebar.dataframe(
                pd.DataFrame(historial_db_data),
                use_container_width=True,
                hide_index=True,
            )

        # MÓDULO 1
        st.markdown("---")
        st.markdown("### 🔌 Módulo 1: Conexión de Balanza y Lectura Rápida de Pallets (SSCC)")
        col_b1, col_b2 = st.columns(2)
        with col_b1:
            st.info("⚡ **Balanza en línea:** el peso se escribe en la última fila de la columna PESO del archivo cargado.")
            peso_capturado = st.number_input("Peso Neto capturado desde Balanza (kg):", value=4.5, step=0.1)
            if st.button("📥 Registrar Peso en Última Fila"):
                df_nuevo, ok_peso, col_peso_usada = funciones.registrar_peso_ultima_fila(df_original, peso_capturado)
                if ok_peso:
                    st.session_state["df_trabajo"] = df_nuevo
                    df_original = df_nuevo
                    funciones.guardar_cambio_db(
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        len(df_original) - 1,
                        col_peso_usada,
                        "(balanza)",
                        str(peso_capturado),
                        auditor_nombre,
                    )
                    st.success(f"Peso de {peso_capturado} kg registrado en la última fila (columna `{col_peso_usada}`).")
                    st.rerun()
                else:
                    st.error("No se encontró una columna de PESO en el archivo, o el archivo está vacío.")
        with col_b2:
            st.info("🔫 **Lector GS1-128 / SSCC:** busca el código en el archivo cargado (CAJA, SSCC, PALLET, CODIGO, LOTE).")
            sscc_input = st.text_input("Escanear Código SSCC o Caja:", placeholder="Ej: 077512345678901234")
            if sscc_input:
                df_sscc = funciones.buscar_registros_por_codigo(df_original, sscc_input)
                if df_sscc.empty:
                    st.error(f"Código `{sscc_input}` no encontrado en la base cargada.")
                else:
                    st.success(f"Unidad encontrada: **{sscc_input}** ({len(df_sscc)} registro(s))")
                    st.dataframe(df_sscc, width="stretch", hide_index=True)

        # MÓDULO 2
        st.markdown("---")
        st.markdown("### 🧪 Módulo 2: Control de Límites Máximos de Residuos (LMR) y Certificación SENASA")
        lotes_disponibles = []
        if cols_traza["lote"]:
            lotes_disponibles = sorted(
                [x for x in df_original[cols_traza["lote"]].astype(str).str.strip().unique() if x and x != "nan"]
            )
        lote_default = lotes_disponibles[0] if lotes_disponibles else ""
        col_mle1, col_mle2, col_mle3 = st.columns(3)
        with col_mle1:
            if lotes_disponibles:
                lote_lmr_sel = st.selectbox("Lote a consultar LMR:", options=lotes_disponibles, index=0)
            else:
                lote_lmr_sel = st.text_input("Ingrese Lote a Consultar LMR:", value=lote_default, placeholder="Ej: LOTE-001")
        with col_mle2:
            analisis_lab = st.selectbox(
                "Resultado de laboratorio (manual / respaldo):",
                ["Usar dato del archivo", "Conforme (Bajo LMR)", "Alerta (Cercano al Límite)", "Rechazado (Supera LMR)"],
            )
        with col_mle3:
            estado_lmr = "sin_dato"
            detalle_lmr = "Sin lote consultado"
            df_lote_lmr = pd.DataFrame()
            if lote_lmr_sel:
                df_lote_lmr = funciones.buscar_registros_por_codigo(df_original, lote_lmr_sel)
                if cols_traza["lote"] and df_lote_lmr.empty:
                    mask_lote = df_original[cols_traza["lote"]].astype(str).str.strip().str.fullmatch(
                        str(lote_lmr_sel).strip(), case=False, na=False
                    )
                    df_lote_lmr = df_original.loc[mask_lote].copy()

                if analisis_lab != "Usar dato del archivo":
                    if "Conforme" in analisis_lab:
                        estado_lmr = "conforme"
                    elif "Alerta" in analisis_lab:
                        estado_lmr = "alerta"
                    else:
                        estado_lmr = "rechazado"
                    detalle_lmr = analisis_lab
                elif not df_lote_lmr.empty and cols_traza["lmr"]:
                    valores_lmr = df_lote_lmr[cols_traza["lmr"]].astype(str).str.strip()
                    estados = [funciones.interpretar_estado_lmr(v) for v in valores_lmr if v and v != "nan"]
                    if "rechazado" in estados:
                        estado_lmr = "rechazado"
                    elif "alerta" in estados:
                        estado_lmr = "alerta"
                    elif "conforme" in estados:
                        estado_lmr = "conforme"
                    detalle_lmr = ", ".join(sorted(set(valores_lmr.head(5))))
                elif df_lote_lmr.empty:
                    detalle_lmr = "Lote no encontrado en archivo"
                else:
                    detalle_lmr = "Lote hallado, sin columna LMR; use el selector manual"

            if estado_lmr == "conforme":
                st.success(f"🟢 **SENASA / Destino:** APROBADO — {detalle_lmr}")
            elif estado_lmr == "alerta":
                st.warning(f"🟡 **SENASA / Destino:** EN CUARENTENA — {detalle_lmr}")
            elif estado_lmr == "rechazado":
                st.error(f"🔴 **SENASA / Destino:** BLOQUEADO DE PLANTA — {detalle_lmr}")
            else:
                st.info(f"ℹ️ **SENASA / Destino:** Sin veredicto automático — {detalle_lmr}")

        if lote_lmr_sel and not df_lote_lmr.empty:
            with st.expander(f"Registros del lote `{lote_lmr_sel}` ({len(df_lote_lmr)})"):
                st.dataframe(df_lote_lmr, width="stretch", hide_index=True)

        # MÓDULO 3
        st.markdown("---")
        st.markdown("### 🗺️ Módulo 3: Trazabilidad Inversa (De Caja o Pallet al Fundo de Origen)")
        col_inv1, col_inv2 = st.columns([2, 3])
        with col_inv1:
            caja_busqueda_inversa = st.text_input(
                "🔍 Ingrese ID de Caja o Pallet para Trazabilidad Inversa:",
                placeholder="Ej: CJ-9842 / SSCC / LOTE",
            )
            cols_detectadas = [f"`{v}` ({k})" for k, v in cols_traza.items() if v]
            if cols_detectadas:
                st.caption("Columnas detectadas: " + ", ".join(cols_detectadas[:8]))
            else:
                st.caption("No se detectaron columnas típicas de trazabilidad; se buscará en todo el archivo.")
        with col_inv2:
            if caja_busqueda_inversa:
                df_traza = funciones.buscar_registros_por_codigo(df_original, caja_busqueda_inversa)
                if df_traza.empty:
                    st.error(
                        f"No se encontró trazabilidad para `{caja_busqueda_inversa}`. "
                        "Verifique que el ID exista en columnas CAJA, PALLET, SSCC, CODIGO o LOTE."
                    )
                else:
                    arbol = funciones.armar_arbol_trazabilidad(
                        df_traza.iloc[0], cols_traza, auditor_nombre, caja_busqueda_inversa
                    )
                    st.markdown(
                        f"""
**Árbol genealógico de trazabilidad para `{arbol['codigo']}`**

- **Fundo / Parcela:** {arbol['fundo']}
- **Productor registrado:** {arbol['productor']}
- **Lote de proceso:** {arbol['lote']}
- **Fecha de cosecha:** {arbol['cosecha']}
- **Turno / línea de packing:** {arbol['turno']}
- **Peso / Calibre:** {arbol['peso']} / {arbol['calibre']}
- **Estado LMR en archivo:** {arbol['lmr']}
- **Inspector responsable:** {arbol['inspector']}
- **Coincidencias:** {len(df_traza)} registro(s)
"""
                    )
                    if len(df_traza) > 1:
                        st.dataframe(df_traza, width="stretch", hide_index=True)

        # MÓDULO 4
        st.markdown("---")
        st.markdown("### 🌡️ Módulo 4: Control Térmico de Cadena de Frío (Pre-frío y Contenedores)")
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            camara_sel = st.selectbox(
                "Cámara Frigorífica / Túnel:",
                ["Pre-Cámara 01", "Túnel de Enfriamiento 03", "Cámara de Almacenamiento 05", "Contenedor Reefer Puerto"],
            )
        with col_f2:
            temp_actual_camara = st.number_input("Temperatura Registrada (°C):", value=temp_min_limite, step=0.5)
        with col_f3:
            frio_ok = temp_actual_camara <= temp_min_limite + 1.0
            estado_frio = "FRÍO ÓPTIMO" if frio_ok else "RUPTURA CADENA DE FRÍO"
            if frio_ok:
                st.success(f"❄️ **Frío Óptimo:** En rango seguro ({temp_actual_camara}°C)")
            else:
                st.error(
                    f"🚨 **Ruptura de Cadena de Frío:** Temperatura alta ({temp_actual_camara}°C). "
                    "Registre la lectura para dejar evidencia en auditoría."
                )
            if st.button("💾 Registrar lectura de frío"):
                funciones.guardar_frio_db(
                    camara_sel,
                    float(temp_actual_camara),
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    estado_frio,
                    auditor_nombre,
                )
                st.success("Lectura de temperatura guardada en SQLite.")
                st.rerun()

        historial_frio = funciones.cargar_frio_db()
        if historial_frio:
            with st.expander("Historial de control de frío (SQLite)"):
                st.dataframe(pd.DataFrame(historial_frio), width="stretch", hide_index=True)

        # MÓDULO 5
        st.markdown("---")
        st.markdown("### 🚢 Módulo 5: Gestión de Contenedores, Bookings y Precintos de Aduanas")
        col_cnt1, col_cnt2, col_cnt3 = st.columns(3)
        with col_cnt1:
            booking_input = st.text_input("Número de Booking:", placeholder="Ej: BKG-998231")
            contenedor_input = st.text_input("ID de Contenedor Reefer:", placeholder="Ej: TGBU1234567")
        with col_cnt2:
            precinto_linea_input = st.text_input("Precinto de Línea Naviera:", placeholder="Ej: MSC-L99821")
            precinto_senasa_input = st.text_input("Precinto Oficial SENASA:", placeholder="Ej: SENASA-004821")
        with col_cnt3:
            st.write("###")
            if st.button("🔒 Registrar y Sellar Contenedor"):
                if booking_input and contenedor_input and precinto_linea_input:
                    funciones.guardar_contenedor_db(booking_input, contenedor_input, precinto_linea_input, precinto_senasa_input, mercado_destino, "SELLADO Y LISTO")
                    st.success(f"¡Contenedor `{contenedor_input}` sellado y registrado en BD con éxito!")
                else:
                    st.error("⚠️ Complete los campos obligatorios de Booking, Contenedor y Precinto.")

        lista_cont_db = funciones.cargar_contenedores_db()
        if lista_cont_db:
            with st.expander("📦 Ver Contenedores Registrados para Despacho"):
                st.dataframe(pd.DataFrame(lista_cont_db), use_container_width=True)

        st.markdown("---")
        st.markdown("### 1️⃣ Selección de Columnas para Exportar")
        todas_cols = list(df_original.columns)

        default_cols = [
            c for c in todas_cols if any(k in c.upper() for k in ["CAJA", "PESO", "LOTE", "CODIGO", "OBServaciones"])
        ]
        if not default_cols:
            default_cols = todas_cols[: min(4, len(todas_cols))]

        cols_elegidas = st.multiselect(
            "Selecciona las columnas que formarán parte del Packing List y Reporte Final:",
            options=todas_cols,
            default=default_cols,
        )

        df_export = df_original[cols_elegidas].copy() if cols_elegidas else df_original.copy()

        st.markdown("### 2️⃣ Limpieza, Auditoría y Reporte Exclusivo de Errores")
        col_l1, col_l2, col_l3, col_l4 = st.columns([1, 1, 1, 1])
        with col_l1:
            if not st.session_state["lote_congelado"]:
                if st.button("🧹 Limpiar Espacios Ocultos"):
                    df_original = df_original.map(lambda x: str(x).strip() if pd.notna(x) else "")
                    df_export = df_original[cols_elegidas].copy() if cols_elegidas else df_original.copy()
                    st.success("¡Espacios eliminados!")
        with col_l2:
            if not st.session_state["lote_congelado"]:
                if st.button("📝 Rellenar Vacíos con (-)"):
                    df_original = df_original.replace("", "-")
                    df_export = df_original[cols_elegidas].copy() if cols_elegidas else df_original.copy()
                    st.success("¡Vacíos reemplazados!")
        with col_l3:
            if st.button("🔍 Ver Registros con Vacíos"):
                st.session_state["mostrar_vacios"] = not st.session_state["mostrar_vacios"]
        with col_l4:
            mask_errores_gen = (df_original == "").any(axis=1)
            col_peso_chk_aux = [c for c in df_original.columns if "PESO" in c.upper()]
            if col_peso_chk_aux:
                pesos_num_aux = pd.to_numeric(df_original[col_peso_chk_aux[0]], errors="coerce")
                mask_errores_gen = mask_errores_gen | (pesos_num_aux < peso_min_caja)
            
            df_solo_errores_dl = df_original[mask_errores_gen]
            pdf_errores_buffer = funciones.generar_pdf_errores(df_solo_errores_dl, archivo.name, producto_sel, auditor_nombre)
            
            st.download_button(
                label="📥 Descargar Solo Errores (PDF)",
                data=pdf_errores_buffer.getvalue(),
                file_name="Reporte_Errores_Planta.pdf",
                mime="application/pdf",
            )

        if st.session_state["mostrar_vacios"]:
            st.warning("⚠️ **Mostrando únicamente filas que contienen campos vacíos o incompletos:**")
            mask_vacios = (df_original == "").any(axis=1)
            df_solo_vacios = df_original[mask_vacios]
            if not df_solo_vacios.empty:
                st.dataframe(df_solo_vacios.style.map(funciones.resaltar_errores_celdas), use_container_width=True)
            else:
                st.success("¡Excelente! No hay registros con celdas vacías en este archivo.")

        st.markdown("### 3️⃣ Control de Calidad, Estadísticas y Alertas")

        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        col_pallet_chk = [c for c in df_original.columns if "PALLET" in c.upper()]
        total_pallets = df_original[col_pallet_chk[0]].nunique() if col_pallet_chk else "N/D"
        
        col_lote_chk = [c for c in df_original.columns if "LOTE" in c.upper()]
        total_lotes = df_original[col_lote_chk[0]].nunique() if col_lote_chk else "N/D"

        col_prod_chk = [c for c in df_original.columns if "PRODUCTOR" in c.upper()]
        total_productores = df_original[col_prod_chk[0]].nunique() if col_prod_chk else "N/D"

        with col_m1:
            st.metric("Total Pallets (SSCC)", total_pallets)
        with col_m2:
            st.metric("Lotes en Proceso", total_lotes)
        with col_m3:
            st.metric("Productores Implicados", total_productores)
        with col_m4:
            st.metric("Duplicados Detectados", total_duplicados)

        col_peso_chk = [c for c in df_original.columns if "PESO" in c.upper()]
        if col_peso_chk:
            pesos_num = pd.to_numeric(df_original[col_peso_chk[0]], errors="coerce")
            cajas_livianas = (pesos_num < peso_min_caja).sum()
            peso_promedio = round(pesos_num.mean(), 2) if not pesos_num.empty else 0
            peso_std = round(pesos_num.std(), 2) if not pesos_num.empty else 0
            st.info(f"📊 **Estadística de Pesaje:** Peso Promedio del Lote: **{peso_promedio} kg** | Desviación Estándar: **{peso_std} kg**")

            if cajas_livianas > 0:
                st.error(f"⚠️ **Alerta Crítica de Pesaje:** Se encontraron {cajas_livianas} registros por debajo del peso mínimo de {peso_min_caja} kg.")

        col1, col2 = st.columns(2)
        with col1:
            col_calibre = [c for c in df_original.columns if "CALIBRE" in c.upper()]
            if col_calibre:
                conteo_calibres = df_original[col_calibre[0]].value_counts().reset_index()
                conteo_calibres.columns = ["Calibre", "Cajas"]
                st.write("📊 **Distribución de Calibres (Tabla):**")
                st.dataframe(conteo_calibres.T, use_container_width=True)
                st.write("📈 **Gráfico de Calibres:**")
                st.bar_chart(conteo_calibres.set_index("Calibre"))

        with col2:
            col_cat = [c for c in df_original.columns if "CATEGORIA" in c.upper() or "CAT" in c.upper()]
            if col_cat:
                conteo_cat = df_original[col_cat[0]].value_counts()
                descarte = conteo_cat.get("DESCARTE", 0) + conteo_cat.get("MERMA", 0)
                porcentaje_merma = round((descarte / total_filas) * 100, 2)

                if porcentaje_merma > max_merma_permitida:
                    st.error(f"🚨 **Alerta de Merma Elevada:** {porcentaje_merma}% de descarte (Límite: {max_merma_permitida}%).")
                else:
                    st.success(f"✅ **Merma Bajo Control:** {porcentaje_merma}% de descarte.")

        st.markdown("---")
        st.markdown("### 🛡️ Módulo de Seguridad: Sello Criptográfico ECC del Lote")
        resumen_datos = f"Auditoría de Planta - Cultivo: {producto_sel} - Fecha: {datetime.date.today()} - Registros: {total_filas}"

        try:
            llave_publica, sello_digital = motor_planta.firmar_reporte_ecc(resumen_datos)
            backend = motor_planta.motor_activo()
            if backend == "rust":
                st.success("🔒 Reporte asegurado con ECC P-256 (motor Rust nativo)")
            else:
                st.success("🔒 Reporte asegurado con ECC P-256 (motor Python / Streamlit Cloud)")
            st.code(f"Sello Digital (Firma ECC): {sello_digital}")
            st.caption(f"Llave Pública de Verificación: {llave_publica}")
            st.caption(f"Backend criptográfico activo: `{backend}`")
        except Exception as e:
            llave_publica = "LLAVE_NO_DISPONIBLE"
            sello_digital = "SELLO_NO_DISPONIBLE"
            st.error(f"No se pudo generar el sello criptográfico: {e}")

        st.markdown("### 4️⃣ Observaciones de Turno, Edición y Cierre")
        col_bus1, col_bus2, col_bus3 = st.columns(3)
        with col_bus1:
            texto_busqueda = st.text_input("🔍 Búsqueda Global (Lote/Contenedor):")
        with col_bus2:
            if col_prod_chk:
                lista_prod = ["TODOS"] + list(df_original[col_prod_chk[0]].unique())
                prod_filtro_sel = st.selectbox("🎯 Filtrar por Productor:", lista_prod)
            else:
                prod_filtro_sel = "TODOS"
        with col_bus3:
            turno_sel = st.selectbox("🕒 Filtrar por Turno de Trabajo:", ["TODOS", "Mañana (06:00 - 14:00)", "Tarde (14:00 - 22:00)", "Noche (22:00 - 06:00)"])

        df_mostrar = df_export
        if texto_busqueda.strip() != "":
            mask_busq = df_mostrar.astype(str).apply(
                lambda row: row.str.contains(texto_busqueda, case=False, na=False).any(),
                axis=1,
            )
            df_mostrar = df_mostrar[mask_busq]

        if col_prod_chk and prod_filtro_sel != "TODOS":
            mask_prod = df_original[col_prod_chk[0]] == prod_filtro_sel
            indices_validos = df_original[mask_prod].index
            df_mostrar = df_mostrar.loc[df_mostrar.index.isin(indices_validos)]

        TAMANO_PAGINA = 100
        total_registros_visibles = len(df_mostrar)
        
        if total_registros_visibles > TAMANO_PAGINA:
            total_paginas = (total_registros_visibles // TAMANO_PAGINA) + (1 if total_registros_visibles % TAMANO_PAGINA > 0 else 0)
            pagina_actual = st.number_input("Página de visualización:", min_value=1, max_value=total_paginas, step=1)
            inicio = (pagina_actual - 1) * TAMANO_PAGINA
            fin = inicio + TAMANO_PAGINA
            df_paginado = df_mostrar.iloc[inicio:fin]
            st.caption(f"Mostrando registros {inicio + 1} al {min(fin, total_registros_visibles)} de un total de {total_registros_visibles} filtrados.")
        else:
            df_paginado = df_mostrar

        if st.session_state["lote_congelado"]:
            st.warning("🔒 **Lote Congelado:** La tabla está en modo solo lectura. Para editar, desbloquee el lote abajo.")
            df_editado = df_mostrar
        else:
            st.info("💡 **Nota de Planta:** Puedes escribir directamente en la columna `Observaciones_Rechazo` de la tabla si una caja presenta golpes, calibre fuera de rango u otra incidencia.")
            df_editado_pag = st.data_editor(df_paginado, use_container_width=True, key="editor_datos_pag")
            
            df_editado = df_mostrar.copy()
            if total_registros_visibles > TAMANO_PAGINA:
                df_editado.iloc[inicio:fin] = df_editado_pag
            else:
                df_editado = df_editado_pag

            if not df_mostrar.equals(df_editado):
                for i in range(len(df_mostrar)):
                    for col in df_mostrar.columns:
                        val_orig = df_mostrar.iloc[i][col]
                        val_nuevo = df_editado.iloc[i][col]
                        if val_orig != val_nuevo:
                            fecha_hora_cambio = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            funciones.guardar_cambio_db(fecha_hora_cambio, i, col, val_orig, val_nuevo, auditor_nombre)

        st.markdown("#### ✅ Verificaciones Fitosanitarias Obligatorias")
        chk_pulpa = st.checkbox("Se verificó la temperatura de pulpa y los límites máximos de residuos (LMR).")
        chk_cuerpo = st.checkbox("El lote se encuentra libre de materias extrañas y plagas cuarentenarias exigidas por SENASA.")
        
        capa_texto = st.text_area("📝 Registro General de Acciones Correctivas (CAPA) / Resumen de Incidencias:", placeholder="Escribir observaciones generales de turno...")

        col_c1, col_c2 = st.columns(2)
        with col_c1:
            if not st.session_state["lote_congelado"]:
                if st.button("🔒 Congelar y Aprobar Lote (Cierre Oficial)"):
                    if chk_pulpa and chk_cuerpo:
                        st.session_state["lote_congelado"] = True
                        st.success("¡Lote aprobado, firmado y congelado correctamente para gerencia!")
                        st.rerun()
                    else:
                        st.error("⚠️ Debe marcar todas las verificaciones del checklist obligatorio antes de aprobar el lote.")
            else:
                if st.button("🔓 Descongelar Lote (Habilitar Edición)"):
                    st.session_state["lote_congelado"] = False
                    st.warning("Lote habilitado para modificaciones.")
                    st.rerun()

        bitacora_db_data = funciones.cargar_bitacora_db()
        if bitacora_db_data:
            with st.expander("📜 Ver Bitácora de Auditoría Persistente (SQLite Audit Trail)"):
                st.dataframe(pd.DataFrame(bitacora_db_data), use_container_width=True)

        st.markdown("### 5️⃣ Exportación de Reportes Oficiales")
        col_ex1, col_ex2, col_ex3 = st.columns(3)

        with col_ex1:
            buffer_completo = io.BytesIO()
            with pd.ExcelWriter(buffer_completo, engine="openpyxl") as writer:
                df_editado.to_excel(writer, index=False, sheet_name="Packing_List")

                pd.DataFrame({
                    "Parámetro de Control": [
                        "Fecha de Emisión",
                        "Cultivo Procesado",
                        "Mercado Destino",
                        "Total Registros Exportados",
                        "Errores Iniciales Detectados",
                        "Confiabilidad del Proceso (Motor Rust)",
                        "Estado del Lote",
                        "Inspector Responsable",
                    ],
                    "Detalle": [
                        datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        producto_sel,
                        mercado_destino,
                        total_filas,
                        total_errores,
                        f"{porcentaje_limpio}%",
                        "CONGELADO / APROBADO" if st.session_state["lote_congelado"] else "EN REVISIÓN",
                        auditor_nombre,
                    ],
                }).to_excel(writer, index=False, sheet_name="Trazabilidad_Resumen")

                if bitacora_db_data:
                    pd.DataFrame(bitacora_db_data).to_excel(writer, index=False, sheet_name="Audit_Trail")

                for sheetname in writer.sheets:
                    worksheet = writer.sheets[sheetname]
                    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    for row in worksheet.iter_rows(min_row=2):
                        for cell in row:
                            if cell.value == "" or cell.value == "-":
                                cell.fill = red_fill

                    for col in worksheet.columns:
                        max_len = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            try:
                                if cell.value:
                                    max_len = max(max_len, len(str(cell.value)))
                            except:
                                pass
                        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

            st.download_button(
                label="📥 Descargar Excel con Hojas Múltiples",
                data=buffer_completo.getvalue(),
                file_name="Reporte_Completo_Exportacion.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        with col_ex2:
            pdf_buffer = funciones.generar_pdf_resumen(
                archivo.name,
                total_filas,
                total_errores,
                total_duplicados,
                porcentaje_limpio,
                producto_sel,
                auditor_nombre,
                st.session_state["lote_congelado"],
                mercado_destino,
                capa_texto,
                sello_digital,
                llave_publica,
            )
            st.download_button(
                label="📄 Descargar PDF Ejecutivo Firmado",
                data=pdf_buffer.getvalue(),
                file_name="Resumen_Ejecutivo_Firmado_ECC.pdf",
                mime="application/pdf",
            )

        with col_ex3:
            buffer_packing = io.BytesIO()
            df_editado.to_csv(buffer_packing, index=False)
            st.download_button(
                label="🚢 Descargar Packing List (CSV)",
                data=buffer_packing.getvalue(),
                file_name="Packing_List_Oficial.csv",
                mime="text/csv",
            )

    except Exception as e:
        st.error(f"Ocurrió un error al procesar el archivo: {e}")